#!/usr/bin/env python3
from __future__ import annotations
"""
auto_detect_layer.py

Analyze PyTorch profiler traces to discover model layer structure and kernel breakdown.

Three-step analysis:
  Step 1: Parse cuda-graph-ON trace → kernel statistics (count, sum, avg, percentage)
  Step 2: Parse cuda-graph-OFF trace → layer structure (types, sub-modules)
  Step 3: Combine → per-layer kernel breakdown with sub-module labels

Usage:
    # Full analysis (both traces)
    python auto_detect_layer.py --graph-on on.trace.json.gz --graph-off off.trace.json.gz

    # Step 1 only: kernel statistics
    python auto_detect_layer.py --graph-on on.trace.json.gz

    # Step 2 only: layer structure
    python auto_detect_layer.py --graph-off off.trace.json.gz

    # Export step 1 to CSV
    python auto_detect_layer.py --graph-on on.trace.json.gz --csv kernels.csv
"""

import argparse
import bisect
import csv
import gzip
import json
import re
import sys
from collections import Counter, OrderedDict
from pathlib import Path


# ---------------------------------------------------------------------------
# Loader
# ---------------------------------------------------------------------------

def load_trace(path: str) -> dict | list:
    p = Path(path)
    if not p.exists():
        sys.exit(f"[ERROR] File not found: {path}")
    opener = gzip.open if path.endswith(".gz") else open
    with opener(path, "rt", encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# GPU kernel extraction
# ---------------------------------------------------------------------------

def extract_gpu_kernels(trace: dict | list, stream: int | None = None) -> list[dict]:
    """Extract GPU kernel events (cat='kernel' only, no memcpy/memset)."""
    events = trace if isinstance(trace, list) else trace.get("traceEvents", [])
    kernels = []
    for ev in events:
        if not isinstance(ev, dict) or ev.get("ph") != "X":
            continue
        if ev.get("cat", "").lower() != "kernel":
            continue
        if stream is not None and ev.get("tid") != stream:
            continue
        ts = float(ev.get("ts", 0))
        dur = float(ev.get("dur", 0))
        kernels.append({
            "name": ev.get("name", "<unknown>"),
            "ts": ts, "ts_end": ts + dur, "dur": dur,
            "cat": ev.get("cat", ""),
            "pid": ev.get("pid", ""), "tid": ev.get("tid", ""),
            "args": ev.get("args", {}),
        })
    kernels.sort(key=lambda k: k["ts"])
    return kernels


def auto_detect_stream(trace: dict | list,
                        single_stream_threshold: float = 0.8) -> int | None:
    """Return a single dominant stream id, or None if kernels are spread
    across many streams (e.g. GLM5 on B200 where each layer runs on its
    own stream).  Passing stream=None to extract_gpu_kernels collects all
    streams, which is the correct behaviour for multi-stream models.
    """
    events = trace if isinstance(trace, list) else trace.get("traceEvents", [])
    counts: Counter = Counter()
    for ev in events:
        if not isinstance(ev, dict) or ev.get("ph") != "X":
            continue
        if ev.get("cat", "").lower() != "kernel":
            continue
        counts[ev.get("tid")] += 1
    if not counts:
        sys.exit("[ERROR] No GPU kernels found in trace.")
    total = sum(counts.values())
    top_stream, top_count = counts.most_common(1)[0]
    if top_count / total >= single_stream_threshold:
        return top_stream
    print(f"[INFO] Kernels spread across {len(counts)} streams "
          f"(top stream has {top_count/total:.0%} of {total} kernels); "
          f"collecting ALL streams.", file=sys.stderr)
    return None


def fmt_dur(us: float) -> str:
    if us >= 1_000_000:
        return f"{us/1_000_000:.3f} s"
    if us >= 1_000:
        return f"{us/1_000:.3f} ms"
    return f"{us:.1f} us"


# ===================================================================
# Step 1: Kernel statistics from cuda-graph-ON trace
# ===================================================================

def compute_kernel_stats(kernels: list[dict]) -> list[dict]:
    """Aggregate kernel statistics: count, sum, avg, percentage."""
    stats: dict[str, dict] = {}
    total_dur = 0.0
    for k in kernels:
        name = k["name"]
        if name not in stats:
            stats[name] = {"name": name, "count": 0, "sum_dur": 0.0}
        stats[name]["count"] += 1
        stats[name]["sum_dur"] += k["dur"]
        total_dur += k["dur"]

    result = []
    for s in stats.values():
        s["avg_dur"] = s["sum_dur"] / s["count"] if s["count"] > 0 else 0
        s["pct"] = s["sum_dur"] / total_dur * 100 if total_dur > 0 else 0
        result.append(s)

    result.sort(key=lambda s: s["sum_dur"], reverse=True)
    return result


def print_step1(stats: list[dict]) -> None:
    total_dur = sum(s["sum_dur"] for s in stats)
    total_count = sum(s["count"] for s in stats)

    print(f"\n{'='*120}")
    print(f" Step 1: Kernel Statistics (cuda-graph-ON)")
    print(f" Total: {total_count} kernel calls, {fmt_dur(total_dur)}")
    print(f"{'='*120}")
    print(f"  {'Kernel Name':<80s} {'Count':>6} {'Sum(us)':>10} {'Avg(us)':>10} {'Pct':>6}")
    print(f"  {'-'*80} {'-'*6} {'-'*10} {'-'*10} {'-'*6}")

    for s in stats:
        name = s["name"][:80]
        print(f"  {name:<80s} {s['count']:>6} {s['sum_dur']:>10.1f} "
              f"{s['avg_dur']:>10.1f} {s['pct']:>5.1f}%")

    print(f"{'='*120}\n")


def _write_xlsx(headers: list[str], rows: list[list], path: str) -> None:
    """Write data to Excel with Arial font and auto-width columns."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active

    arial = Font(name="Arial", size=10)
    arial_bold = Font(name="Arial", size=10, bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                              fill_type="solid")

    for c, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=val)
        cell.font = arial_bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    for r, row_data in enumerate(rows, 2):
        for c, val in enumerate(row_data, 1):
            if isinstance(val, str):
                try:
                    val = float(val) if val and "." in val and val.replace(".", "").replace("-", "").isdigit() else val
                except (ValueError, TypeError):
                    pass
            cell = ws.cell(row=r, column=c, value=val if val != "" else None)
            cell.font = arial

    for c in range(1, len(headers) + 1):
        max_len = len(str(headers[c - 1]))
        for r in range(2, min(len(rows) + 2, 50)):
            val = ws.cell(row=r, column=c).value
            if val:
                max_len = max(max_len, min(len(str(val)), 60))
        ws.column_dimensions[get_column_letter(c)].width = max_len + 2

    wb.save(path)


def write_step1(stats: list[dict], path: str) -> None:
    headers = ["Name", "Count", "SumDuration_us", "AvgDuration_us", "Percentage"]
    rows = [[s["name"], s["count"], round(s["sum_dur"], 1),
             round(s["avg_dur"], 3), round(s["pct"], 2)] for s in stats]
    _write_xlsx(headers, rows, path)
    print(f"[INFO] Step 1 written to: {path}", file=sys.stderr)


def _strip_jit_hash(name: str) -> str:
    """Strip JIT/Triton compilation hashes from kernel names.

    JIT-compiled kernels (Triton, tilelang) often embed a run-specific
    hex hash in their name so the same kernel may appear as:
      graph-OFF:  act_quant_kernel_abc12345__kernel
      graph-ON:   act_quant_kernel_def67890__kernel
    Stripping the hash allows cross-trace name matching.
    """
    return re.sub(r'_[0-9a-fA-F]{6,}(?=_|$)', '', name)



def _lookup_stat(name: str,
                 stat_lookup: dict, stat_lookup_norm: dict) -> tuple[dict | None, str]:
    """Three-level kernel stat lookup; returns (stat_dict_or_None, method_str)."""
    s = stat_lookup.get(name)
    if s is not None:
        return s, "exact"
    s = stat_lookup_norm.get(_strip_jit_hash(name))
    if s is not None:
        return s, "norm"
    return None, ""


def write_step3(layer_types: list[dict], kernels_off: list[dict],
                kernel_stats: list[dict] | None, path: str,
                callsite_map: dict | None = None) -> None:
    """Export step 3 breakdown to Excel.

    Structure comes from graph-OFF trace (layer types, kernel order, sections,
    modules, call sites).  When graph-ON stats are available and a kernel name
    can be matched (exact or hash-stripped), the graph-ON kernel name and
    avg_dur are used as primary values; the graph-OFF name and single-pass
    duration are kept as reference columns.

    Output columns:
      LayerType, LayerCount, Index, Section, LeafModule,
      KernelName,       ← graph-ON name if matched, else graph-OFF name
      AvgDuration_us,   ← graph-ON avg_dur if matched, else graph-OFF dur
      Count, SumDuration_us, Percentage,  ← graph-ON stats (empty if not matched)
      MatchMethod,      ← "exact" / "norm" / "none"
      GraphOFF_KernelName, GraphOFF_Duration_us,  ← graph-OFF reference
      CallSite
    """
    # Build stat lookups (exact + hash-normalized)
    stat_lookup: dict[str, dict] = {}
    stat_lookup_norm: dict[str, dict] = {}
    if kernel_stats:
        for s in kernel_stats:
            stat_lookup[s["name"]] = s
            norm = _strip_jit_hash(s["name"])
            if norm not in stat_lookup_norm:
                stat_lookup_norm[norm] = s

    headers = ["LayerType", "LayerCount", "Index", "Section", "LeafModule",
               "KernelName", "AvgDuration_us",
               "Count", "SumDuration_us", "Percentage",
               "MatchMethod",
               "GraphOFF_KernelName", "GraphOFF_Duration_us",
               "CallSite"]

    rows = []
    for i, lt in enumerate(layer_types):
        if i > 0:
            rows.append([""] * len(headers))
        label = chr(ord("A") + i)
        sub_mod_str = " + ".join(lt["sub_modules"])
        for pos, (kidx, top_mod, leaf) in enumerate(lt.get("kernel_breakdown", [])):
            k = kernels_off[kidx]
            off_name = k["name"]
            off_dur = round(k["dur"], 1)
            cs = (callsite_map or {}).get(kidx, "")

            if kernel_stats:
                s, method = _lookup_stat(off_name, stat_lookup, stat_lookup_norm)
            else:
                s, method = None, "none"

            if s:
                kernel_name = s["name"]
                avg_dur = round(s["avg_dur"], 3)
                count = s["count"]
                sum_dur = round(s["sum_dur"], 1)
                pct = round(s["pct"], 2)
            else:
                kernel_name = off_name
                avg_dur = off_dur
                count = ""
                sum_dur = ""
                pct = ""
                method = "none"

            row = [f"{label}: {sub_mod_str}", lt["count"], pos,
                   top_mod, leaf or "(self)",
                   kernel_name, avg_dur,
                   count, sum_dur, pct,
                   method,
                   off_name, off_dur,
                   cs]
            rows.append(row)

    _write_xlsx(headers, rows, path)
    n_matched = sum(1 for r in rows if r and r[10] != "none" and r[10] != "")
    n_total = sum(1 for r in rows if r and r[2] != "")
    print(f"[INFO] Step 3 written to: {path} "
          f"({n_total} kernels, {n_matched} matched to graph-ON)",
          file=sys.stderr)


# ===================================================================
# Step 2: Layer structure from cuda-graph-OFF trace
# ===================================================================

def find_decoder_layer_class(trace: dict | list) -> str | None:
    """Find the nn.Module class that represents a DecoderLayer."""
    events = trace if isinstance(trace, list) else trace.get("traceEvents", [])
    module_events = [
        ev for ev in events
        if isinstance(ev, dict) and ev.get("ph") == "X"
        and ev.get("name", "").startswith("nn.Module: ") and ev.get("dur", 0) > 0
    ]
    if not module_events:
        return None

    class_info: dict[str, dict] = {}
    for ev in module_events:
        cls = re.sub(r'_\d+$', '', ev["name"].replace("nn.Module: ", ""))
        if cls not in class_info:
            class_info[cls] = {"instances": set(), "total_dur": 0, "count": 0}
        class_info[cls]["instances"].add(ev["name"])
        class_info[cls]["total_dur"] += ev.get("dur", 0)
        class_info[cls]["count"] += 1

    candidates = [(cls, len(info["instances"]), info["total_dur"] / info["count"])
                  for cls, info in class_info.items() if len(info["instances"]) >= 3]
    if not candidates:
        return None

    named = [c for c in candidates if "Decoder" in c[0] or "Layer" in c[0]]
    pool = named if named else candidates
    return max(pool, key=lambda c: c[1] * c[2])[0]


def get_one_forward_pass(trace: dict | list, cls_name: str) -> list[dict]:
    """Get nn.Module events for one forward pass (2nd pass to skip warmup)."""
    events = trace if isinstance(trace, list) else trace.get("traceEvents", [])
    all_events = sorted(
        [ev for ev in events
         if isinstance(ev, dict) and ev.get("ph") == "X"
         and re.sub(r'_\d+$', '', ev.get("name", "").replace("nn.Module: ", "")) == cls_name],
        key=lambda ev: ev["ts"],
    )
    instances = set(ev["name"] for ev in all_events)
    n_layers = len(instances)
    if n_layers == 0:
        return []
    n_passes = len(all_events) // n_layers
    pass_idx = min(1, n_passes - 1)  # 2nd pass if available
    return all_events[pass_idx * n_layers: (pass_idx + 1) * n_layers]


def find_direct_children(trace: dict | list, parent_ev: dict,
                          target_classes: set[str] | None = None) -> list[dict]:
    """Find direct nn.Module children of a parent module event."""
    events = trace if isinstance(trace, list) else trace.get("traceEvents", [])
    pts, pend = parent_ev["ts"], parent_ev["ts"] + parent_ev.get("dur", 0)

    # Get all nn.Module events within parent's time range (excluding parent)
    children = []
    for ev in events:
        if not isinstance(ev, dict) or ev.get("ph") != "X":
            continue
        if not ev.get("name", "").startswith("nn.Module: "):
            continue
        if ev is parent_ev or ev["name"] == parent_ev["name"]:
            continue
        ets = ev.get("ts", 0)
        eend = ets + ev.get("dur", 0)
        if ets >= pts and eend <= pend + 1:
            cls = re.sub(r'_\d+$', '', ev["name"].replace("nn.Module: ", ""))
            if target_classes is None or cls in target_classes:
                children.append(ev)

    # Filter to direct children only (not nested inside another child)
    children.sort(key=lambda e: e["ts"])
    direct = []
    for c in children:
        cts, cend = c["ts"], c["ts"] + c.get("dur", 0)
        is_nested = any(
            o["ts"] < cts and o["ts"] + o.get("dur", 0) > cend
            for o in children if o is not c
        )
        if not is_nested:
            direct.append(c)

    return direct


def build_ext_id_map(trace: dict | list, kernels: list[dict]) -> tuple:
    """Build kernel lookup maps and sorted runtime events.

    Returns (ext_to_kidx, corr_to_kidx, runtime, rt_ts) where:
      ext_to_kidx: External id → kernel index
      corr_to_kidx: correlation → kernel index (for kernels without External id)
      runtime: sorted cuda_runtime events (with External id or correlation)
      rt_ts: timestamps for binary search
    """
    events = trace if isinstance(trace, list) else trace.get("traceEvents", [])
    ext_to_kidx = {}
    corr_to_kidx = {}
    for i, k in enumerate(kernels):
        eid = k.get("args", {}).get("External id")
        corr = k.get("args", {}).get("correlation")
        if eid is not None and eid not in ext_to_kidx:
            ext_to_kidx[eid] = i
        elif eid is None and corr is not None and corr not in corr_to_kidx:
            corr_to_kidx[corr] = i

    runtime = sorted(
        [ev for ev in events
         if isinstance(ev, dict) and ev.get("cat") == "cuda_runtime"
         and ev.get("ph") == "X"
         and (ev.get("args", {}).get("External id") is not None
              or ev.get("args", {}).get("correlation") is not None)],
        key=lambda ev: ev["ts"],
    )
    rt_ts = [ev["ts"] for ev in runtime]
    return ext_to_kidx, corr_to_kidx, runtime, rt_ts


def build_callsite_map(trace: dict | list, kernels: list[dict],
                        ext_to_kidx: dict, runtime: list[dict],
                        target_kidxs: set[int] | None = None) -> dict:
    """
    Build kernel_index → call site string mapping.
    Finds the innermost python_function from sglang source code
    that contains each kernel's cuda_runtime launch event.

    If target_kidxs is given, only resolve those kernel indices (fast).
    """
    events = trace if isinstance(trace, list) else trace.get("traceEvents", [])

    # Build per-thread sorted list of python_function events
    py_by_thread: dict[int, list] = {}
    py_ts_by_thread: dict[int, list] = {}
    for ev in events:
        if not isinstance(ev, dict) or ev.get("cat") != "python_function":
            continue
        if ev.get("ph") != "X" or ".py" not in ev.get("name", ""):
            continue
        tid = ev.get("tid")
        if tid not in py_by_thread:
            py_by_thread[tid] = []
        py_by_thread[tid].append(ev)

    for tid in py_by_thread:
        py_by_thread[tid].sort(key=lambda e: e["ts"])
        py_ts_by_thread[tid] = [e["ts"] for e in py_by_thread[tid]]

    # Build ext_id and correlation → runtime event mappings
    ext_to_runtime = {}
    corr_to_runtime = {}
    for rt in runtime:
        args = rt.get("args", {})
        eid = args.get("External id")
        corr = args.get("correlation")
        if eid is not None and eid not in ext_to_runtime:
            ext_to_runtime[eid] = rt
        if corr is not None and corr not in corr_to_runtime:
            corr_to_runtime[corr] = rt

    # Resolve call sites
    callsite_map = {}
    resolve_list = target_kidxs if target_kidxs else range(len(kernels))

    for i in resolve_list:
        k = kernels[i]
        args = k.get("args", {})
        eid = args.get("External id")
        rt = ext_to_runtime.get(eid) if eid is not None else None
        if rt is None:
            corr = args.get("correlation")
            rt = corr_to_runtime.get(corr) if corr is not None else None
        if rt is None:
            continue
        rt_ts_val = rt["ts"]
        rt_tid = rt["tid"]

        if rt_tid not in py_by_thread:
            continue

        py_list = py_by_thread[rt_tid]
        ts_list = py_ts_by_thread[rt_tid]

        # Binary search to find approximate position
        pos = bisect.bisect_right(ts_list, rt_ts_val) - 1

        # Scan backwards. Don't break on gaps — outer functions with
        # longer duration may have started earlier and still contain rt_ts.
        # Limit scan to 2000 events or 50ms back to keep it fast.
        best = None
        best_dur = float("inf")
        scan_limit = max(0, pos - 2000)
        ts_limit = rt_ts_val - 50_000  # 50ms back max

        for j in range(max(0, pos), scan_limit - 1, -1):
            pf = py_list[j]
            pf_ts = pf["ts"]
            if pf_ts > rt_ts_val:
                continue
            if pf_ts < ts_limit:
                break  # too far back in time
            pf_end = pf_ts + pf.get("dur", 0)
            if pf_end < rt_ts_val:
                continue  # this one doesn't contain rt, but keep scanning
            # This event contains the runtime call
            name = pf["name"]
            dur = pf.get("dur", 0)
            if "/sglang/" in name and dur < best_dur:
                best = name
                best_dur = dur

        if best:
            callsite_map[i] = best

    return callsite_map


def find_layer_sections(pf_events: list[dict], layer_ev: dict) -> list[dict]:
    """Find section-level functions/modules within a decoder layer.

    Sections are direct children of the decoder layer's forward method,
    such as prepare_attn, DeepseekV2AttentionMLA, prepare_mlp, etc.

    Args:
        pf_events: python_function events within the layer's time range,
                   sorted by (ts, -dur).
        layer_ev: the decoder layer nn.Module event.

    Returns sorted list of {"name": str, "ts": float, "end": float}
    """
    if not pf_events:
        return []

    layer_dur = layer_ev.get("dur", 1)

    # Find the model's forward function (sglang source, large duration)
    fwd = None
    for ev in pf_events:
        name = ev.get("name", "")
        if ("/sglang/" in name and ": forward" in name
                and ev.get("dur", 0) > layer_dur * 0.5):
            fwd = ev
            break

    if fwd is None:
        return []

    fts = fwd["ts"]
    fend = fts + fwd.get("dur", 0)

    # Collect events contained in forward
    children = [ev for ev in pf_events
                if ev is not fwd
                and ev["ts"] >= fts
                and ev["ts"] + ev.get("dur", 0) <= fend + 1]

    # Keep only direct children (not nested inside another child)
    children.sort(key=lambda e: e["ts"])
    direct = []
    for c in children:
        cts, cend = c["ts"], c["ts"] + c.get("dur", 0)
        if not any(o["ts"] < cts and o["ts"] + o.get("dur", 0) > cend
                   for o in children if o is not c):
            direct.append(c)

    # Filter to significant functions and extract clean names
    sections = []
    for c in direct:
        name = c.get("name", "")
        dur = c.get("dur", 0)
        # Skip tiny helpers and torch internals
        if dur < 1:
            continue
        if ("torch/" in name or "nn/modules/" in name
                or "<built-in" in name) and not name.startswith("nn.Module:"):
            continue

        if name.startswith("nn.Module: "):
            clean = re.sub(r'_\d+$', '', name.replace("nn.Module: ", ""))
        else:
            match = re.search(r':\s*(\w+)\s*$', name)
            clean = match.group(1) if match else name

        sections.append({"name": clean, "ts": c["ts"],
                         "end": c["ts"] + dur})

    sections.sort(key=lambda s: s["ts"])
    return sections


def find_pf_leaf(all_pf: list[dict], all_pf_ts: list[float],
                 rt_ts: float, section_name: str) -> str:
    """Find innermost python_function as fallback leaf label.

    Used when no nn.Module wraps a kernel. Looks for the innermost
    non-torch, non-module python_function that contains the runtime
    event, excluding the section function itself.

    Returns "filename.py(line)" or empty string.
    """
    pos = bisect.bisect_right(all_pf_ts, rt_ts) - 1
    best_name = None
    best_dur = float("inf")

    for j in range(max(0, pos), max(0, pos - 2000), -1):
        pf = all_pf[j]
        pf_ts = pf["ts"]
        if pf_ts > rt_ts:
            continue
        if pf_ts < rt_ts - 50_000:
            break
        pf_end = pf_ts + pf.get("dur", 0)
        if pf_end < rt_ts:
            continue
        name = pf.get("name", "")
        dur = pf.get("dur", 0)
        # Skip torch internals, nn.Module wrappers, built-ins, triton runtime
        if ("torch/" in name or "nn/modules/" in name
                or "<built-in" in name or name.startswith("nn.Module:")
                or "triton/runtime/" in name or "triton/backends/" in name):
            continue
        if dur < best_dur:
            best_name = name
            best_dur = dur

    if not best_name:
        return ""

    # Extract "filename.py(line)" from full path
    m = re.search(r'([^/]+\.py\(\d+\))', best_name)
    if not m:
        return ""

    short = m.group(1)
    # Skip if it's the section's own function (e.g. forward, prepare_attn)
    func_match = re.search(r':\s*(\w+)\s*$', best_name)
    if func_match and func_match.group(1) == section_name:
        return ""

    return short


def get_section_name(sections: list[dict], ts: float) -> str:
    """Find which section contains the given timestamp."""
    for s in sections:
        if s["ts"] <= ts <= s["end"]:
            return s["name"]
    return "(layer)"


def get_kernels_for_module(module_ev: dict, ext_to_kidx: dict,
                            corr_to_kidx: dict,
                            runtime: list[dict], rt_ts: list[float]) -> list[int]:
    """Get GPU kernel indices launched by a module event."""
    mod_ts = module_ev["ts"]
    mod_end = mod_ts + module_ev.get("dur", 0)
    lo = bisect.bisect_left(rt_ts, mod_ts)
    hi = bisect.bisect_right(rt_ts, mod_end)
    kidxs = set()
    for j in range(lo, hi):
        args = runtime[j].get("args", {})
        eid = args.get("External id")
        if eid is not None and eid in ext_to_kidx:
            kidxs.add(ext_to_kidx[eid])
        else:
            corr = args.get("correlation")
            if corr is not None and corr in corr_to_kidx:
                kidxs.add(corr_to_kidx[corr])
    return sorted(kidxs)


def find_deep_kernel_labels(module_events: list[dict], layer_ev: dict,
                             kidxs: list[int], kernels: list[dict],
                             ext_to_rt: dict, corr_to_rt: dict) -> dict:
    """Find deepest nn.Module path for each kernel within a layer.

    Returns {kernel_idx: "ParentModule > LeafModule"}.
    """
    pts = layer_ev["ts"]
    pend = pts + layer_ev.get("dur", 0)

    # Collect all nn.Module descendants within layer's time range
    descendants = []
    for ev in module_events:
        if ev is layer_ev:
            continue
        ets = ev.get("ts", 0)
        eend = ets + ev.get("dur", 0)
        if ets >= pts and eend <= pend + 1:
            cls = re.sub(r'_\d+$', '', ev["name"].replace("nn.Module: ", ""))
            descendants.append({
                "cls": cls, "ts": ets, "end": eend, "dur": ev.get("dur", 0),
            })

    if not descendants:
        return {}

    # Build hierarchy paths: sort by duration desc (parents first)
    descendants.sort(key=lambda m: m["dur"], reverse=True)
    for i, m in enumerate(descendants):
        parent_path = None
        parent_dur = float("inf")
        for j in range(i):
            p = descendants[j]
            if (p["ts"] <= m["ts"] and p["end"] >= m["end"] - 1
                    and p["dur"] < parent_dur):
                parent_path = p["path"]
                parent_dur = p["dur"]
        m["path"] = f"{parent_path} > {m['cls']}" if parent_path else m["cls"]

    # For each kernel, find deepest enclosing module via runtime event timestamp
    labels = {}
    for ki in kidxs:
        k = kernels[ki]
        args = k.get("args", {})
        eid = args.get("External id")
        rt = ext_to_rt.get(eid) if eid is not None else None
        if rt is None:
            corr = args.get("correlation")
            rt = corr_to_rt.get(corr) if corr is not None else None
        if rt is None:
            continue
        rt_ts_val = rt["ts"]

        best = None
        best_dur = float("inf")
        for m in descendants:
            if m["ts"] <= rt_ts_val <= m["end"] and m["dur"] < best_dur:
                best = m
                best_dur = m["dur"]

        if best:
            labels[ki] = best["path"]

    return labels


def analyze_layer_structure(trace: dict | list, kernels: list[dict]):
    """
    Step 2: Discover layer structure.
    Returns (cls_name, layer_types, callsite_map) where layer_types is:
      [ { 'name': str, 'count': int, 'layers': [name, ...],
          'sub_modules': [cls, ...],
          'kernel_breakdown': [ (kernel_idx, top_module, leaf_detail), ... ] }, ... ]
    and callsite_map is { kernel_idx: "file.py(line): func_name" }
    """
    cls_name = find_decoder_layer_class(trace)
    if cls_name is None:
        return None, [], {}

    forward_pass = get_one_forward_pass(trace, cls_name)
    if not forward_pass:
        return cls_name, [], {}

    ext_to_kidx, corr_to_kidx, runtime, rt_ts = build_ext_id_map(trace, kernels)

    # Pre-compute module events and runtime mapping for deep labeling
    events = trace if isinstance(trace, list) else trace.get("traceEvents", [])
    module_events = [
        ev for ev in events
        if isinstance(ev, dict) and ev.get("ph") == "X"
        and ev.get("name", "").startswith("nn.Module: ")
    ]
    module_events.sort(key=lambda e: e["ts"])

    ext_to_rt = {}
    corr_to_rt = {}
    for rt in runtime:
        args = rt.get("args", {})
        eid = args.get("External id")
        corr = args.get("correlation")
        if eid is not None and eid not in ext_to_rt:
            ext_to_rt[eid] = rt
        if corr is not None and corr not in corr_to_rt:
            corr_to_rt[corr] = rt

    # Discover significant child classes by sampling multiple layers
    # (different layer types may have different children)
    significant_classes = set()
    sample_indices = set()
    # Sample a few layers from different positions
    for idx in [0, 1, len(forward_pass) // 2, len(forward_pass) - 1]:
        sample_indices.add(min(idx, len(forward_pass) - 1))
    for idx in sample_indices:
        sample = forward_pass[idx]
        all_children = find_direct_children(trace, sample)
        for c in all_children:
            cls = re.sub(r'_\d+$', '', c["name"].replace("nn.Module: ", ""))
            avg_dur = c.get("dur", 0)
            if avg_dur > sample.get("dur", 1) * 0.05:
                significant_classes.add(cls)

    print(f"[INFO] Found nn.Module: {cls_name} "
          f"({len(forward_pass)} layers per forward pass)", file=sys.stderr)
    print(f"[INFO] Sub-module classes: {', '.join(sorted(significant_classes))}",
          file=sys.stderr)

    # Pre-filter python_function events for section detection
    all_pf = sorted(
        [ev for ev in events
         if isinstance(ev, dict) and ev.get("ph") == "X"
         and ev.get("cat") == "python_function" and ev.get("dur", 0) > 0],
        key=lambda e: (e["ts"], -e.get("dur", 0)),
    )
    all_pf_ts = [e["ts"] for e in all_pf]

    # Analyze each layer
    layer_data = []
    for layer_ev in forward_pass:
        layer_name = layer_ev["name"].replace("nn.Module: ", "")
        children = find_direct_children(trace, layer_ev, significant_classes)
        sub_mods = [re.sub(r'_\d+$', '', c["name"].replace("nn.Module: ", ""))
                    for c in children]

        # Find sections (high-level code blocks in the layer's forward)
        lts = layer_ev["ts"]
        lend = lts + layer_ev.get("dur", 0)
        pf_lo = bisect.bisect_left(all_pf_ts, lts)
        pf_hi = bisect.bisect_right(all_pf_ts, lend)
        sections = find_layer_sections(all_pf[pf_lo:pf_hi], layer_ev)

        # Get kernel-to-module mapping (deep — find leaf module)
        all_kidxs = get_kernels_for_module(layer_ev, ext_to_kidx,
                                            corr_to_kidx, runtime, rt_ts)
        kernel_labels = find_deep_kernel_labels(
            module_events, layer_ev, all_kidxs, kernels,
            ext_to_rt, corr_to_rt)

        breakdown = []
        for ki in all_kidxs:
            # Determine section from python_function context
            k_args = kernels[ki].get("args", {})
            eid = k_args.get("External id")
            rt = ext_to_rt.get(eid) if eid is not None else None
            if rt is None:
                corr = k_args.get("correlation")
                rt = corr_to_rt.get(corr) if corr is not None else None
            section = get_section_name(sections, rt["ts"]) if rt else "(layer)"

            # Leaf module from nn.Module hierarchy
            leaf = kernel_labels.get(ki, "")
            # Strip section prefix from leaf if it starts with the section name
            if leaf.startswith(section + " > "):
                leaf = leaf[len(section) + 3:]
            elif leaf == section:
                leaf = ""

            # Fallback: use innermost python_function when no nn.Module
            if not leaf and rt:
                leaf = find_pf_leaf(all_pf, all_pf_ts, rt["ts"], section)

            breakdown.append((ki, section, leaf))
        layer_data.append({
            "name": layer_name,
            "sub_modules": tuple(sub_mods),
            "kernel_breakdown": breakdown,
            "n_kernels": len(all_kidxs),
        })

    # === Fallback: recover unlinked kernels via GPU-timestamp interpolation ===
    # Some kernels (e.g. launched via cudaLaunchKernelExC) lack matching
    # cuda_runtime events, so get_kernels_for_module cannot find them.
    # We handle two sub-cases:
    #   (a) Kernels that fall WITHIN the forward-pass GPU time window →
    #       assign to the same layer/section as the nearest preceding linked kernel.
    #   (b) Kernels that fall OUTSIDE (after) the forward-pass window →
    #       these are inter-layer / model-level ops (allreduce, residual norms, etc.)
    #       that execute between decoder layers.  Collect them into a synthetic
    #       "(inter-layer)" layer_data entry so they show up labelled in Step 3.
    linked_entries = []  # (gpu_ts, kernel_idx, layer_data_idx, section)
    for ldi, ld in enumerate(layer_data):
        for ki, section, _leaf in ld["kernel_breakdown"]:
            linked_entries.append((kernels[ki]["ts"], ki, ldi, section))
    linked_entries.sort()

    if linked_entries:
        assigned_kidxs = {e[1] for e in linked_entries}
        fp_gpu_start = linked_entries[0][0]
        fp_gpu_end = max(kernels[e[1]]["ts_end"] for e in linked_entries)
        entry_ts = [e[0] for e in linked_entries]

        unlinked_by_layer: dict[int, list] = {}
        inter_layer_kidxs: list[int] = []   # case (b): outside forward-pass window

        for ki in range(len(kernels)):
            if ki in assigned_kidxs:
                continue
            k = kernels[ki]
            if k["ts"] < fp_gpu_start:
                continue  # before forward pass — ignore
            if k["ts"] > fp_gpu_end:
                # case (b): inter-layer / post-layer kernel
                inter_layer_kidxs.append(ki)
                continue
            # case (a): within window — assign to nearest preceding linked kernel
            pos = bisect.bisect_right(entry_ts, k["ts"]) - 1
            if pos < 0:
                ldi = linked_entries[0][2]
                inferred_section = linked_entries[0][3]
            else:
                ldi = linked_entries[pos][2]
                inferred_section = linked_entries[pos][3]
            unlinked_by_layer.setdefault(ldi, []).append(
                (ki, inferred_section))

        n_recovered = 0
        for ldi, items in unlinked_by_layer.items():
            ld = layer_data[ldi]
            existing = list(ld["kernel_breakdown"])
            for ki, section in items:
                existing.append((ki, section, ""))
            existing.sort(key=lambda item: kernels[item[0]]["ts"])
            ld["kernel_breakdown"] = existing
            n_recovered += len(items)

        if n_recovered:
            print(f"[INFO] Recovered {n_recovered} unlinked kernels via "
                  f"GPU-timestamp interpolation", file=sys.stderr)

        if inter_layer_kidxs:
            print(f"[INFO] Skipping {len(inter_layer_kidxs)} kernels outside "
                  f"decoder-layer GPU window (inter-layer / post-layer ops)",
                  file=sys.stderr)

    # Group by sub_module signature
    groups: OrderedDict = OrderedDict()
    for i, ld in enumerate(layer_data):
        key = ld["sub_modules"]
        if key not in groups:
            groups[key] = {"layers": [], "indices": [], "example_breakdown": None}
        groups[key]["layers"].append(ld["name"])
        groups[key]["indices"].append(i)
        if groups[key]["example_breakdown"] is None and i >= 1:
            groups[key]["example_breakdown"] = ld["kernel_breakdown"]
    # Fallback: use first if no 2nd instance
    for key, g in groups.items():
        if g["example_breakdown"] is None:
            idx = g["indices"][0]
            g["example_breakdown"] = layer_data[idx]["kernel_breakdown"]

    layer_types = []
    for sub_mods, g in groups.items():
        layer_types.append({
            "sub_modules": list(sub_mods),
            "count": len(g["layers"]),
            "layers": g["layers"],
            "indices": g["indices"],
            "kernel_breakdown": g["example_breakdown"],
        })

    # Build callsite map only for kernels in breakdowns (fast)
    target_kidxs = set()
    for lt in layer_types:
        for item in lt.get("kernel_breakdown", []):
            target_kidxs.add(item[0])
    print(f"[INFO] Resolving call sites for {len(target_kidxs)} kernels...",
          file=sys.stderr)
    callsite_map = build_callsite_map(trace, kernels, ext_to_kidx, runtime,
                                      target_kidxs)

    return cls_name, layer_types, callsite_map


def print_step2(cls_name: str, layer_types: list[dict]) -> None:
    total_layers = sum(lt["count"] for lt in layer_types)
    print(f"\n{'='*100}")
    print(f" Step 2: Layer Structure (cuda-graph-OFF)")
    print(f" Model Layer: {cls_name}")
    print(f" Layers per forward pass: {total_layers}")
    print(f" Distinct layer types: {len(layer_types)}")
    print(f"{'='*100}")

    for i, lt in enumerate(layer_types):
        label = chr(ord("A") + i)
        layers = lt["layers"]
        if len(layers) <= 5:
            name_str = ", ".join(layers)
        else:
            name_str = f"{layers[0]} .. {layers[-1]}"

        print(f"\n  Type {label}: {lt['count']} layers")
        print(f"  Layers: {name_str}")
        print(f"  Sub-modules: {' + '.join(lt['sub_modules'])}")

    print(f"\n{'='*100}\n")


# ===================================================================
# Step 3: Per-layer kernel breakdown with sub-module labels
# ===================================================================

def print_step3(cls_name: str, layer_types: list[dict],
                kernels: list[dict], kernel_stats: list[dict] | None,
                callsite_map: dict | None = None) -> None:
    """
    Print kernel breakdown for each layer type.
    Uses graph-OFF kernel data with sub-module labels.
    If kernel_stats (from graph-ON step 1) is provided, also shows
    the graph-ON avg duration for cross-reference.
    """
    # Build name->avg lookup from step 1
    stat_lookup = {}
    if kernel_stats:
        for s in kernel_stats:
            stat_lookup[s["name"]] = s

    print(f"\n{'='*100}")
    print(f" Step 3: Layer Kernel Breakdown")
    print(f"{'='*100}")

    for i, lt in enumerate(layer_types):
        label = chr(ord("A") + i)
        breakdown = lt["kernel_breakdown"]
        if not breakdown:
            continue

        print(f"\n  Type {label} ({lt['count']} layers): "
              f"{' + '.join(lt['sub_modules'])}")
        print(f"  {'-'*90}")

        current_top = None
        top_dur = 0.0
        total_dur = 0.0

        header = f"  {'#':>3}  {'Detail':<40s}  {'Duration':>10}  {'Kernel Name'}"
        if kernel_stats:
            header += f"  {'(graph-ON avg)':>14}"
        print(header)
        print(f"  {'-'*3}  {'-'*40}  {'-'*10}  {'-'*50}")

        for pos, (kidx, top_mod, leaf) in enumerate(breakdown):
            k = kernels[kidx]
            dur = k["dur"]
            total_dur += dur

            if top_mod != current_top:
                if current_top is not None:
                    print(f"  {'':>3}  {'Subtotal':>40}  {fmt_dur(top_dur):>10}")
                    print()
                current_top = top_mod
                top_dur = 0.0
                print(f"  ---- {top_mod} ----")

            top_dur += dur
            detail = leaf if leaf else "(self)"
            name_display = k["name"][:60]
            cs = callsite_map.get(kidx, "") if callsite_map else ""
            line = f"  {pos:>3}  {detail:<40s}  {fmt_dur(dur):>10}  {name_display}"

            if kernel_stats and k["name"] in stat_lookup:
                avg = stat_lookup[k["name"]]["avg_dur"]
                line += f"  {fmt_dur(avg):>14}"

            print(line)
            if cs:
                print(f"  {'':>3}  {'':>40}  {'':>10}  caller: {cs}")

        # Last module subtotal
        if current_top is not None:
            print(f"  {'':>3}  {'Subtotal':>40}  {fmt_dur(top_dur):>10}")

        print(f"\n  {'':>3}  {'TOTAL':>40}  {fmt_dur(total_dur):>10}  "
              f"({len(breakdown)} kernels)")

    print(f"\n{'='*100}\n")


# ===================================================================
# CLI
# ===================================================================

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Analyze PyTorch profiler traces for layer structure and kernel breakdown.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("--graph-on", metavar="TRACE",
                   help="Cuda-graph-ON trace (for kernel statistics)")
    p.add_argument("--graph-off", metavar="TRACE",
                   help="Cuda-graph-OFF trace (for layer structure)")
    p.add_argument("--stream", type=int, default=None,
                   help="GPU stream ID (default: auto-detect)")
    p.add_argument("--out", metavar="DIR",
                   help="Export Excel files to directory (step1_kernel_stats.xlsx, step3_layer_breakdown.xlsx)")
    p.add_argument("--tag", metavar="TAG", default="",
                   help="Append tag to output filenames, e.g. --tag _ATOM → step3_layer_breakdown_ATOM.xlsx")
    return p


def main() -> None:
    args = build_parser().parse_args()
    if not args.graph_on and not args.graph_off:
        sys.exit("Provide at least one: --graph-on or --graph-off")

    # Prepare output dir
    out_dir = None
    if args.out:
        out_dir = Path(args.out)
        out_dir.mkdir(parents=True, exist_ok=True)

    kernel_stats = None
    cls_name = None
    layer_types = None
    kernels_off = None

    # --- Step 1: Kernel statistics from graph-ON trace ---
    if args.graph_on:
        print(f"[INFO] Step 1: Loading graph-ON trace: {args.graph_on}", file=sys.stderr)
        trace_on = load_trace(args.graph_on)
        stream_on = args.stream if args.stream is not None else auto_detect_stream(trace_on)
        kernels_on = extract_gpu_kernels(trace_on, stream=stream_on)
        stream_desc = f"stream {stream_on}" if stream_on is not None else "all streams"
        print(f"[INFO] Found {len(kernels_on)} kernels on {stream_desc}",
              file=sys.stderr)

        kernel_stats = compute_kernel_stats(kernels_on)
        print_step1(kernel_stats)
        del trace_on, kernels_on  # free memory; stats are all we need

        if out_dir:
            write_step1(kernel_stats, str(out_dir / f"step1_kernel_stats{args.tag}.xlsx"))

    # --- Step 2: Layer structure from graph-OFF trace ---
    if args.graph_off:
        print(f"[INFO] Step 2: Loading graph-OFF trace: {args.graph_off}",
              file=sys.stderr)
        trace_off = load_trace(args.graph_off)
        stream_off = args.stream if args.stream is not None else auto_detect_stream(trace_off)
        kernels_off = extract_gpu_kernels(trace_off, stream=stream_off)
        stream_desc = f"stream {stream_off}" if stream_off is not None else "all streams"
        print(f"[INFO] Found {len(kernels_off)} kernels on {stream_desc}",
              file=sys.stderr)

        cls_name, layer_types, callsite_map = analyze_layer_structure(
            trace_off, kernels_off)
        if cls_name:
            print_step2(cls_name, layer_types)

            # --- Step 3: Combined breakdown ---
            if layer_types:
                print_step3(cls_name, layer_types, kernels_off, kernel_stats,
                            callsite_map)
                if out_dir:
                    write_step3(layer_types, kernels_off, kernel_stats,
                                str(out_dir / f"step3_layer_breakdown{args.tag}.xlsx"),
                                callsite_map)
        else:
            print("[WARN] No nn.Module DecoderLayer events found.", file=sys.stderr)


if __name__ == "__main__":
    main()

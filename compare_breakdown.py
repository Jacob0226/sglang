#!/usr/bin/env python3
"""
compare_breakdown.py

Compare two step3_layer_breakdown.csv files (e.g. MI355X vs B200) side by side.
Aligns rows by (ParentModule, LeafModule) using LCS sequence alignment,
preserving execution order and padding the shorter side with empty rows.

Usage:
    python compare_breakdown.py --file-a MI355X/step3_layer_breakdown.xlsx --file-b B200/step3_layer_breakdown.xlsx --labels MI355X B200 --out comparison.xlsx
"""

import argparse
import csv  # for reading input CSVs
import os
import sys


def _trace_busy(path: str):
    """Compute Σduration, union-busy, overlap factor and decode-step count from
    a raw *.trace.json.gz (graph-ON) trace.

    union-busy = wall time the GPU is actually busy (overlapping kernels on
    different streams counted once); overlap = Σduration / union-busy.
    Returns a dict or None on failure.
    """
    import gzip
    import json
    from collections import Counter
    try:
        with gzip.open(path, "rt") as f:
            t = json.load(f)
    except Exception as e:  # noqa: BLE001
        print(f"[warn] could not read trace {path}: {e}", file=sys.stderr)
        return None
    evs = t if isinstance(t, list) else t.get("traceEvents", [])
    k = [e for e in evs if isinstance(e, dict)
         and str(e.get("cat", "")).lower() == "kernel" and e.get("ph") == "X"]
    if not k:
        return None
    iv = sorted((e["ts"], e["ts"] + e["dur"]) for e in k)
    sumd = sum(e["dur"] for e in k)
    tot = 0.0
    cs, ce = iv[0]
    for s, e in iv[1:]:
        if s > ce:
            tot += ce - cs
            cs, ce = s, e
        else:
            ce = max(ce, e)
    tot += ce - cs
    steps = (sum(c for n, c in Counter(e["name"] for e in k).items()
                 if "topk_transform_decode" in n) // 78) or 1
    return {"sum_ms": sumd / 1e3, "busy_ms": tot / 1e3,
            "overlap": (sumd / tot if tot else 0.0),
            "steps": steps, "busy_per_step": tot / steps / 1e3,
            "nkernels": len(k)}


def load_breakdown(path: str) -> list[dict]:
    """Load a step3_layer_breakdown file (.xlsx or .csv), skipping blank rows."""
    rows = []
    if path.endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row_cells in ws.iter_rows(min_row=2):
            vals = [cell.value for cell in row_cells]
            row = {h: (str(v) if v is not None else "") for h, v in zip(headers, vals)}
            if row.get("KernelName"):
                rows.append(row)
        wb.close()
    else:
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get("KernelName"):
                    rows.append(row)
    return rows


# LeafModule aliases: different names that should match during alignment
LEAF_ALIASES = {
    "FlashInferFusedMoE": "FusedMoE/FlashInferFusedMoE",
    "FusedMoE": "FusedMoE/FlashInferFusedMoE",
}


def normalize_leaf(leaf: str) -> str:
    """Normalize LeafModule for cross-platform matching."""
    return LEAF_ALIASES.get(leaf, leaf)


def build_blocks(rows: list[dict]) -> list[tuple]:
    """Build ordered list of ((section, normalized_leaf), [rows...]) blocks.

    Consecutive rows with the same (Section, LeafModule) form one block.
    Non-consecutive duplicates stay separate (e.g. two RadixAttention calls
    represent different attention operations).
    """
    blocks = []
    for row in rows:
        key = (row.get("Section", ""),
               normalize_leaf(row.get("LeafModule", "")))
        if blocks and blocks[-1][0] == key:
            blocks[-1][1].append(row)
        else:
            blocks.append((key, [row]))
    return blocks


def keys_match(ka, kb):
    """Check if two (section, leaf) keys should be considered matching.

    Rules:
    - Exact match after normalization
    - Same section + either side is (self) → flexible match
    """
    sa, la = ka
    sb, lb = kb
    if sa != sb:
        return False
    if la == lb:
        return True
    if la == "(self)" or lb == "(self)":
        return True
    return False


def lcs_alignment(seq_a, seq_b):
    """LCS-based alignment of two key sequences.

    Returns list of (idx_a_or_None, idx_b_or_None) pairs.
    """
    keys_a = [b[0] for b in seq_a]
    keys_b = [b[0] for b in seq_b]
    na, nb = len(keys_a), len(keys_b)

    # Build LCS table using flexible key matching
    dp = [[0] * (nb + 1) for _ in range(na + 1)]
    for i in range(na):
        for j in range(nb):
            if keys_match(keys_a[i], keys_b[j]):
                dp[i + 1][j + 1] = dp[i][j] + 1
            else:
                dp[i + 1][j + 1] = max(dp[i][j + 1], dp[i + 1][j])

    # Backtrack to build alignment
    alignment = []
    i, j = na, nb
    matched = []
    while i > 0 and j > 0:
        if keys_match(keys_a[i - 1], keys_b[j - 1]):
            matched.append((i - 1, j - 1))
            i -= 1
            j -= 1
        elif dp[i - 1][j] >= dp[i][j - 1]:
            i -= 1
        else:
            j -= 1
    matched.reverse()

    # Build full alignment with unmatched items inserted
    ai, bi = 0, 0
    for ma, mb in matched:
        # Insert A-only blocks before this match
        while ai < ma:
            alignment.append((ai, None))
            ai += 1
        # Insert B-only blocks before this match
        while bi < mb:
            alignment.append((None, bi))
            bi += 1
        # Matched pair
        alignment.append((ma, mb))
        ai = ma + 1
        bi = mb + 1

    # Remaining unmatched after last LCS match
    while ai < na:
        alignment.append((ai, None))
        ai += 1
    while bi < nb:
        alignment.append((None, bi))
        bi += 1

    return alignment


def _split_self_blocks(blocks):
    """Split blocks into named blocks and (self) blocks.

    Returns (named_blocks, self_map) where self_map[i] is a list of
    (self) blocks that appear AFTER named_blocks[i].
    Leading (self) blocks go into self_map[-1].
    """
    named = []
    self_map = {-1: []}  # -1 = before any named block
    for key, rows in blocks:
        if key[1] == "(self)":
            idx = len(named) - 1 if named else -1
            self_map.setdefault(idx, []).append((key, rows))
        else:
            named.append((key, rows))
    return named, self_map


def build_comparison(rows_a, rows_b, label_a, label_b):
    """Build aligned comparison rows using LCS.

    Returns (header, output_rows, section_meta) where section_meta is a list
    of dicts with section info for subtotals and summary table.
    """
    blocks_a = build_blocks(rows_a)
    blocks_b = build_blocks(rows_b)

    named_a, self_a = _split_self_blocks(blocks_a)
    named_b, self_b = _split_self_blocks(blocks_b)
    alignment = lcs_alignment(named_a, named_b)

    # --- Consolidate alignment so same (layer_type, section) is not interleaved ---
    # LCS can produce: ...SparseMoeBlock, prepare_mlp, SparseMoeBlock...
    # when unmatched blocks from one side land between matched blocks.
    # Re-order so that all entries of the same section within the same
    # layer type are contiguous.
    def _get_group_key(pair):
        """Return (layer_type, section) for grouping."""
        ia, ib = pair
        if ia is not None:
            rows = named_a[ia][1]
            sec = named_a[ia][0][0]
        else:
            rows = named_b[ib][1]
            sec = named_b[ib][0][0]
        lt = ""
        for r in rows:
            if r.get("LayerType"):
                lt = r["LayerType"]
                break
        return (lt, sec)

    consolidated = []
    for pair in alignment:
        gk = _get_group_key(pair)
        if consolidated and _get_group_key(consolidated[-1][-1]) == gk:
            consolidated[-1].append(pair)
        else:
            # Check if this (layer_type, section) appeared in an earlier group
            merged = False
            for grp in consolidated:
                if _get_group_key(grp[0]) == gk:
                    grp.append(pair)
                    merged = True
                    break
            if not merged:
                consolidated.append([pair])
    alignment = [pair for grp in consolidated for pair in grp]

    header = ["LayerType", "Section", "LeafModule", "LayerCount", "#",
              f"{label_a}_KernelName", f"{label_a}_us",
              f"{label_b}_us", f"{label_b}_KernelName",
              f"{label_a}_CallSite", f"{label_b}_CallSite"]

    empty_row = [""] * len(header)
    output_rows = []
    section_meta = []  # [{section, layer_type, layer_count, data_start, subtotal_idx}]

    cur_section = None
    cur_data_start = 0
    cur_layer_type = ""
    cur_layer_count = ""

    def _emit_self_blocks(idx_a, idx_b):
        for skey, srows in self_a.get(idx_a, []):
            for r in srows:
                output_rows.append(
                    ["", skey[0], skey[1], "", "",
                     r.get("KernelName", ""), r.get("AvgDuration_us", ""),
                     "", "", r.get("CallSite", ""), ""])
        for skey, srows in self_b.get(idx_b, []):
            for r in srows:
                output_rows.append(
                    ["", skey[0], skey[1], "", "",
                     "", "", r.get("AvgDuration_us", ""),
                     r.get("KernelName", ""), "", r.get("CallSite", "")])

    def _close_section():
        nonlocal cur_section
        if cur_section is None:
            return
        # Subtotal row (placeholder — formulas written by write_xlsx)
        output_rows.append(
            ["", cur_section, "Subtotal", "", "", "",
             "__SUM_A__", "__SUM_B__", "", "__RATIO__", ""])
        section_meta.append({
            "section": cur_section,
            "layer_type": cur_layer_type,
            "layer_count": cur_layer_count,
            "data_start": cur_data_start,
            "subtotal_idx": len(output_rows) - 1,
        })
        output_rows.append(empty_row)
        cur_section = None

    _emit_self_blocks(-1, -1)

    for ia, ib in alignment:
        ka = named_a[ia][1] if ia is not None else []
        kb = named_b[ib][1] if ib is not None else []

        key_a_val = named_a[ia][0] if ia is not None else None
        key_b_val = named_b[ib][0] if ib is not None else None
        key = key_a_val or key_b_val
        parent, leaf_a = key
        if key_a_val and key_b_val and key_a_val[1] != key_b_val[1]:
            la, lb = key_a_val[1], key_b_val[1]
            leaf = lb if la == "(self)" else la
        else:
            leaf = leaf_a

        layer_type = ""
        layer_count = ""
        for src in (ka, kb):
            if src:
                if src[0].get("LayerType"):
                    layer_type = src[0]["LayerType"]
                if src[0].get("LayerCount"):
                    layer_count = src[0]["LayerCount"]
                break
        # Coerce LayerCount to int so the data cell stores a number, allowing
        # the Summary's `=D<row>` reference to participate in arithmetic.
        if layer_count:
            try:
                layer_count = int(layer_count)
            except (ValueError, TypeError):
                pass

        # Detect section change → close previous section
        if parent != cur_section:
            _close_section()
            cur_section = parent
            cur_data_start = len(output_rows)
            cur_layer_type = layer_type
            cur_layer_count = layer_count

        n = max(len(ka), len(kb))
        for i in range(n):
            ra = ka[i] if i < len(ka) else {}
            rb = kb[i] if i < len(kb) else {}
            output_rows.append([
                layer_type if i == 0 else "",
                parent if i == 0 else "",
                leaf if i == 0 else "",
                layer_count if i == 0 else "",
                str(i),
                ra.get("KernelName", ""),
                ra.get("AvgDuration_us", ""),
                rb.get("AvgDuration_us", ""),
                rb.get("KernelName", ""),
                ra.get("CallSite", ""),
                rb.get("CallSite", ""),
            ])

        _emit_self_blocks(ia, ib)

    _close_section()  # close last section

    return header, output_rows, section_meta


def write_xlsx(header, rows, section_meta, label_a, label_b, path, meta_info=None):
    """Write comparison to Excel with formulas, subtotals, and summary.

    meta_info (optional dict): adds a metadata block above the table describing
    the profile and the source traces, plus overlap-factor / union-busy stats.
    Keys: title, a_name, b_name, a_busy, b_busy (the *_busy dicts come from
    _trace_busy()).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    arial = Font(name="Arial", size=10)
    arial_bold = Font(name="Arial", size=10, bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                              fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))

    # --- Metadata block (above the table). RO = number of rows it occupies, so
    # every absolute row reference below is shifted down by RO. ---
    RO = 0
    if meta_info:
        meta_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6",
                                fill_type="solid")
        meta_lines = []
        meta_lines.append(f"Profile: {meta_info.get('title', '') or '(unspecified)'}")
        meta_lines.append(f"{label_a} trace: {meta_info.get('a_name', '')}")
        meta_lines.append(f"{label_b} trace: {meta_info.get('b_name', '')}")
        ab = meta_info.get("a_busy")
        bb = meta_info.get("b_busy")
        if ab and bb:
            meta_lines.append(
                f"GPU-busy (graph-ON, per decode step): "
                f"{label_a} {ab['busy_per_step']:.2f} ms (overlap {ab['overlap']:.2f}x) | "
                f"{label_b} {bb['busy_per_step']:.2f} ms (overlap {bb['overlap']:.2f}x) | "
                f"{label_b}/{label_a} = "
                f"{(bb['busy_per_step']/ab['busy_per_step']*100 if ab['busy_per_step'] else 0):.0f}%"
            )
            meta_lines.append(
                "NOTE: the Σduration table/Summary below ignores cross-stream "
                "overlap; use the GPU-busy line above for wall-clock (TPOT) comparison."
            )
        elif ab or bb:
            one = ab or bb
            who = label_a if ab else label_b
            meta_lines.append(
                f"GPU-busy (graph-ON): {who} {one['busy_per_step']:.2f} ms/step "
                f"(overlap {one['overlap']:.2f}x)")
        for i, ln in enumerate(meta_lines, 1):
            c = ws.cell(row=i, column=1, value=ln)
            c.font = arial_bold if i == 1 else arial
            c.fill = meta_fill
        RO = len(meta_lines) + 1  # +1 blank separator row before the table

    # Assign a distinct subtotal color per unique layer_type (in order of first appearance)
    SUBTOTAL_PALETTE = [
        "E2EFDA",  # green
        "FCE4D6",  # orange
        "EAD1DC",  # pink/rose
        "D6DCE4",  # grey-blue
        "E2D9F3",  # lavender
        "D9F0D3",  # mint
        "FFF2CC",  # yellow
    ]
    layer_type_color: dict[str, str] = {}

    def _get_subtotal_fill(layer_type: str) -> PatternFill:
        if layer_type not in layer_type_color:
            idx = len(layer_type_color) % len(SUBTOTAL_PALETTE)
            layer_type_color[layer_type] = SUBTOTAL_PALETTE[idx]
        color = layer_type_color[layer_type]
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Pre-populate color map in section_meta order so colors are stable
    for meta in section_meta:
        _get_subtotal_fill(meta["layer_type"])

    COL_A_US = 7   # column G
    COL_B_US = 8   # column H

    # --- Write header ---
    for c, val in enumerate(header, 1):
        cell = ws.cell(row=1 + RO, column=c, value=val)
        cell.font = arial_bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = f"A{2 + RO}"

    # Build a map from excel row -> layer_type for subtotal rows
    subtotal_layer_type: dict[int, str] = {}
    for meta in section_meta:
        xl_sub = meta["subtotal_idx"] + 2 + RO  # +2: 1-indexed + header
        subtotal_layer_type[xl_sub] = meta["layer_type"]

    # --- Write data rows ---
    for r, row_data in enumerate(rows, 2 + RO):
        is_separator = not any(row_data)
        is_subtotal = any(v in ("__SUM_A__", "__SUM_B__", "__RATIO__")
                          for v in row_data if isinstance(v, str))
        row_fill = _get_subtotal_fill(subtotal_layer_type.get(r, "")) if is_subtotal else None
        for c, val in enumerate(row_data, 1):
            if val in ("__SUM_A__", "__SUM_B__", "__RATIO__"):
                continue  # filled later with formulas
            try:
                val = float(val) if val and isinstance(val, str) and "." in val and val.replace(".", "").replace("-", "").isdigit() else val
            except (ValueError, TypeError):
                pass
            cell = ws.cell(row=r, column=c, value=val if val != "" else None)
            cell.font = arial_bold if is_subtotal else arial
            if is_subtotal and row_fill:
                cell.fill = row_fill
        if is_separator:
            for c in range(1, len(header) + 1):
                ws.cell(row=r, column=c).border = thin_border

    # --- Write subtotal formulas ---
    ca = get_column_letter(COL_A_US)
    cb = get_column_letter(COL_B_US)

    for meta in section_meta:
        xl_row = meta["subtotal_idx"] + 2 + RO  # +2: 1-indexed + header
        data_start_xl = meta["data_start"] + 2 + RO
        data_end_xl = xl_row - 1  # row before subtotal
        sfill = _get_subtotal_fill(meta["layer_type"])

        cell_a = ws.cell(row=xl_row, column=COL_A_US)
        cell_a.value = f"=SUM({ca}{data_start_xl}:{ca}{data_end_xl})"
        cell_a.font = arial_bold
        cell_a.fill = sfill
        cell_a.number_format = "0.000"

        cell_b = ws.cell(row=xl_row, column=COL_B_US)
        cell_b.value = f"=SUM({cb}{data_start_xl}:{cb}{data_end_xl})"
        cell_b.font = arial_bold
        cell_b.fill = sfill
        cell_b.number_format = "0.000"

        # Ratio in column J (10)
        cell_r = ws.cell(row=xl_row, column=10)
        cell_r.value = f"=IF({ca}{xl_row}>0,{cb}{xl_row}/{ca}{xl_row},\"\")"
        cell_r.font = arial_bold
        cell_r.fill = sfill
        cell_r.number_format = "0%"

    # --- Summary table ---
    summary_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC",
                               fill_type="solid")
    all_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))

    summary_start = len(rows) + 4 + RO  # 2 rows gap after data
    r = summary_start

    # --- Optimization columns ---
    opt_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                           fill_type="solid")
    NUM_COLS = 13  # A..M

    # Summary header row 1
    ws.cell(row=r, column=1, value="Summary").font = arial_bold
    ws.cell(row=r, column=1).fill = summary_fill
    for c in range(2, NUM_COLS + 1):
        ws.cell(row=r, column=c).fill = summary_fill
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
    cell = ws.cell(row=r, column=6, value="Dur-1Layer")
    cell.font = arial_bold
    cell.fill = summary_fill
    cell.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)
    cell = ws.cell(row=r, column=8, value="Dur-1Forward")
    cell.font = arial_bold
    cell.fill = summary_fill
    cell.alignment = Alignment(horizontal="center")
    # Optimization header (green)
    ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=13)
    cell = ws.cell(row=r, column=11, value="Optimization and Projection")
    cell.font = arial_bold
    cell.fill = opt_fill
    cell.alignment = Alignment(horizontal="center")
    for c in range(12, 14):
        ws.cell(row=r, column=c).fill = opt_fill
    for c in range(1, NUM_COLS + 1):
        ws.cell(row=r, column=c).border = all_border
    r += 1

    # Summary header row 2
    sum_headers2 = ["LayerType", "Section", "", "#Layer", "",
                    label_a, label_b, label_a, label_b, f"{label_b}/{label_a}",
                    "Section Speedup", "Overall Speedup", "Optimized Perf (Accum)"]
    for c, val in enumerate(sum_headers2, 1):
        cell = ws.cell(row=r, column=c, value=val if val else None)
        cell.font = arial_bold
        cell.fill = opt_fill if c >= 11 else summary_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = all_border
    r += 1

    # Baseline row (optimization columns only)
    baseline_r = r
    ws.cell(row=r, column=11, value="0 (Baseline)").font = arial
    ws.cell(row=r, column=12, value="0 (Baseline)").font = arial
    # M13 baseline filled later (needs total_row reference)
    for c in range(1, NUM_COLS + 1):
        ws.cell(row=r, column=c).border = all_border
    r += 1

    # Summary data rows
    data_first = r
    for meta in section_meta:
        # Reference the source rows in the data table above instead of writing
        # values directly, so edits to the table propagate to the summary.
        # NOTE: use this meta's own subtotal_idx (not a section-name-keyed
        # dict) since the same Section name can appear under multiple
        # LayerType groups (e.g. prepare_attn shows up for both Layer A and B)
        # and a flat dict would let the second occurrence overwrite the first.
        st_row = meta["subtotal_idx"] + 2 + RO
        data_start_xl = meta["data_start"] + 2 + RO

        ws.cell(row=r, column=1,
                value=f"=A{data_start_xl}").font = arial
        ws.cell(row=r, column=2,
                value=f"=B{data_start_xl}").font = arial
        ws.cell(row=r, column=4,
                value=f"=D{data_start_xl}").font = arial

        ws.cell(row=r, column=6,
                value=f"={ca}{st_row}").font = arial
        ws.cell(row=r, column=6).number_format = "0.000"
        ws.cell(row=r, column=7,
                value=f"={cb}{st_row}").font = arial
        ws.cell(row=r, column=7).number_format = "0.000"

        ws.cell(row=r, column=8,
                value=f"=F{r}*D{r}").font = arial
        ws.cell(row=r, column=8).number_format = "0.0"
        ws.cell(row=r, column=9,
                value=f"=G{r}*D{r}").font = arial
        ws.cell(row=r, column=9).number_format = "0.0"

        ws.cell(row=r, column=10,
                value=f"=IF(H{r}>0,I{r}/H{r},\"\")").font = arial
        ws.cell(row=r, column=10).number_format = "0%"

        for c in range(1, NUM_COLS + 1):
            ws.cell(row=r, column=c).border = all_border
        r += 1

    # Total row (no borders)
    total_end = r - 1
    total_row = r
    ws.cell(row=r, column=2, value="Total").font = arial_bold
    for col in (8, 9):
        cl = get_column_letter(col)
        ws.cell(row=r, column=col,
                value=f"=SUM({cl}{data_first}:{cl}{total_end})").font = arial_bold
        ws.cell(row=r, column=col).number_format = "0.0"
    ws.cell(row=r, column=10,
            value=f"=IF(H{r}>0,I{r}/H{r},\"\")").font = arial_bold
    ws.cell(row=r, column=10).number_format = "0%"

    # --- Fill optimization formulas (need total_row) ---
    # Baseline Optimized Perf = B200_Total / MI355X_Total
    c13 = ws.cell(row=baseline_r, column=13)
    c13.value = f"=IF($H${total_row}>0,$I${total_row}/$H${total_row},\"\")"
    c13.font = arial
    c13.number_format = "0%"

    for i, meta in enumerate(section_meta):
        dr = data_first + i

        # Section Speedup = MI355X_section / B200_section (if MI355X slower)
        ws.cell(row=dr, column=11,
                value=f"=IF(H{dr}>I{dr},H{dr}/I{dr},0)").font = arial
        ws.cell(row=dr, column=11).number_format = "0.0"

        # Overall Speedup = ReduceTime / MI355X_Total
        # ReduceTime = IF(H > I, H - I, 0)
        ws.cell(row=dr, column=12,
                value=f"=IF(H{dr}>I{dr},(H{dr}-I{dr})/$H${total_row},0)").font = arial
        ws.cell(row=dr, column=12).number_format = "0.0%"

        # Optimized Perf (Accum) = previous row M + current L
        prev_r = baseline_r if i == 0 else (dr - 1)
        ws.cell(row=dr, column=13,
                value=f"=M{prev_r}+L{dr}").font = arial
        ws.cell(row=dr, column=13).number_format = "0%"

    # --- Auto-fit column widths ---
    for c in range(1, len(header) + 1):
        max_len = len(str(header[c - 1]))
        for row in range(2 + RO, min(len(rows) + 2 + RO, 50 + RO)):
            val = ws.cell(row=row, column=c).value
            if val and not str(val).startswith("="):
                max_len = max(max_len, min(len(str(val)), 60))
        ws.column_dimensions[get_column_letter(c)].width = max_len + 2

    wb.save(path)


def main():
    p = argparse.ArgumentParser(
        description="Compare two step3_layer_breakdown.csv files side by side.")
    p.add_argument("--file-a", required=True, metavar="FILE",
                   help="First breakdown file (.xlsx or .csv)")
    p.add_argument("--file-b", required=True, metavar="FILE",
                   help="Second breakdown file (.xlsx or .csv)")
    p.add_argument("--labels", nargs=2, default=["A", "B"],
                   metavar=("A_LABEL", "B_LABEL"),
                   help="Labels for the two platforms (default: A B)")
    p.add_argument("--out", required=True, metavar="XLSX",
                   help="Output Excel path (e.g. comparison.xlsx)")
    p.add_argument("--title", default="",
                   help="Profile description written in the first row of the output")
    p.add_argument("--trace-a", default="", metavar="TRACE",
                   help="Source graph-ON trace for label A; its name is shown and "
                        "its overlap-factor / union-busy are computed")
    p.add_argument("--trace-b", default="", metavar="TRACE",
                   help="Source graph-ON trace for label B (name + overlap/union-busy)")
    args = p.parse_args()

    rows_a = load_breakdown(args.file_a)
    rows_b = load_breakdown(args.file_b)

    if not rows_a:
        sys.exit(f"[ERROR] No data in {args.file_a}")
    if not rows_b:
        sys.exit(f"[ERROR] No data in {args.file_b}")

    blocks_a = build_blocks(rows_a)
    blocks_b = build_blocks(rows_b)
    print(f"[INFO] {args.labels[0]}: {len(rows_a)} kernels, "
          f"{len(blocks_a)} blocks", file=sys.stderr)
    print(f"[INFO] {args.labels[1]}: {len(rows_b)} kernels, "
          f"{len(blocks_b)} blocks", file=sys.stderr)

    header, output, section_meta = build_comparison(
        rows_a, rows_b, args.labels[0], args.labels[1])

    meta_info = {
        "title": args.title,
        "a_name": os.path.basename(args.trace_a) if args.trace_a else os.path.basename(args.file_a),
        "b_name": os.path.basename(args.trace_b) if args.trace_b else os.path.basename(args.file_b),
        "a_busy": _trace_busy(args.trace_a) if args.trace_a else None,
        "b_busy": _trace_busy(args.trace_b) if args.trace_b else None,
    }

    write_xlsx(header, output, section_meta,
               args.labels[0], args.labels[1], args.out, meta_info)
    print(f"[INFO] Written to {args.out}", file=sys.stderr)


if __name__ == "__main__":
    main()

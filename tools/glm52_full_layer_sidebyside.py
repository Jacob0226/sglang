#!/usr/bin/env python3
"""
GLM-5.2-SPECIALIZED full/shared decode-layer side-by-side (ATOM | SGLang).

Why specialized: SGLang's no-cuda-graph decode trace has NO per-layer
(`model.layers.N`) annotations — only `step[DECODE bs=4]` + torch.compile FX-graph
call markers. So analyze_trace.py groups kernels by *module signature*, which
merges GLM-5.2's MoE-full and MoE-shared layers (identical modules, differing only
by whether the indexer ran) into one group. That makes a clean full-vs-shared
compare impossible via the step3 xlsx.

This tool instead segments a decode step POSITIONALLY using GLM-5.2 structure:
  - `fused_qk_rmsnorm` fires exactly once per layer  -> layer anchor.
  - the `allreduce_fusion` right before it is the layer's prepare_attn boundary.
  - layer i = [allreduce_before(rmsnorm[i]), allreduce_before(rmsnorm[i+1])).
  - MoE layer  = layer index >= first_k_dense_replace (3).
  - full layer = its kernels include the indexer (kn_entry_2c / paged_mqa_logits).
Section labels are assigned positionally (state flips to MoE at the first MoE GEMM).
Per-kernel TIMES come from the graph-ON decode trace (avg per kernel name), same as
the ATOM side; structure/order comes from the no-graph trace.

Usage:
  python glm52_full_layer_sidebyside.py \
    --atom-time A.graph.json.gz --atom-struct A.nograph.json.gz \
    --sglang-time S.graph.json.gz --sglang-struct S.nograph.json.gz \
    --layer-kind full --out out.xlsx
"""
import argparse, gzip, json, bisect, re, collections, sys, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from atom_sglang_layer_sidebyside import (
    load, _k, _g, _dom, graph_name_avg, atom_layer, short, canon,
)

FIRST_K_DENSE = 3


def _sglang_kernel_time_avg(evs):
    """avg duration per kernel name over one median decode step of the graph-on trace."""
    ks = _k(evs); p, t = _dom(ks); ks = [k for k in ks if (k["pid"], k["tid"]) == (p, t)]
    ann = [a for a in _g(evs) if (a["pid"], a["tid"]) == (p, t)]
    steps = sorted([a for a in ann if a["name"].startswith("step[DECODE") or a["name"].startswith("decode[")],
                   key=lambda a: a["ts"])
    if steps:
        s = steps[len(steps) // 2]; s0, s1 = s["ts"], s["ts"] + s["dur"]
        ks = [k for k in ks if s0 <= k["ts"] < s1]
    agg = collections.defaultdict(lambda: [0, 0.0])
    for k in ks:
        agg[k["name"]][0] += 1; agg[k["name"]][1] += k["dur"]
    return {n: v[1] / v[0] for n, v in agg.items()}


def _canon_positional(kn, state):
    """Section for a SGLang kernel using name + running state ('MLA'|'MoE').
    hgemm is ambiguous (MLA q/o-proj vs MoE/dense gate/up/down) so it follows state."""
    k = kn.lower()
    if ("allreduce" in k or "all_reduce" in k or "reduce_1stage" in k
            or "cross_device_reduce" in k):
        return "norm/comm", state
    if "add_rmsnorm_quant" in k:
        return "norm/comm", state
    if any(t in k for t in ("mfma_moe", "moe_sorting", "grouped_topk", "moe_sort",
                            "append_shared_experts")):
        return "MoE/MLP", "MoE"
    if "act_and_mul" in k or ("silu" in k and "moe" not in k):
        return "MoE/MLP", "MoE"
    if "hgemm" in k:
        return ("MoE/MLP" if state == "MoE" else "MLA_attention"), state
    return "MLA_attention", state


def sglang_layer_glm52(struct_evs, timeavg, want_full):
    ks = _k(struct_evs); p, t = _dom(ks); ks = [k for k in ks if (k["pid"], k["tid"]) == (p, t)]
    ks.sort(key=lambda k: k["ts"])
    ann = [a for a in _g(struct_evs) if (a["pid"], a["tid"]) == (p, t)]
    steps = sorted([a for a in ann if a["name"].startswith("step[DECODE")], key=lambda a: a["ts"])
    if not steps:
        raise SystemExit("no step[DECODE ...] annotation in sglang struct trace")
    s = steps[len(steps) // 2]; s0, s1 = s["ts"], s["ts"] + s["dur"]
    sk = [k for k in ks if s0 <= k["ts"] < s1]
    rms = sorted(k["ts"] for k in sk if "fused_qk_rmsnorm" in k["name"])
    ar = sorted(k["ts"] for k in sk if "allreduce_fusion" in k["name"] or "cross_device_reduce" in k["name"])
    if len(rms) < 5:
        raise SystemExit(f"expected ~78 fused_qk_rmsnorm anchors, got {len(rms)}")

    def ar_before(ts):
        i = bisect.bisect_left(ar, ts) - 1
        return ar[i] if i >= 0 else ts

    # layer i spans [ar_before(rms[i]), ar_before(rms[i+1]))
    def layer_span(i):
        a = ar_before(rms[i]); b = ar_before(rms[i + 1]) if i + 1 < len(rms) else s1
        return a, b

    def has_indexer(a, b):
        return any(("paged_mqa_logits" in k["name"] or "kn_entry_2c" in k["name"]
                    or "topk_transform" in k["name"]) and a <= k["ts"] < b for k in sk)

    # pick target layer: MoE (i>=FIRST_K_DENSE) with indexer==want_full
    pick = None
    for i in range(len(rms) - 1):
        if i < FIRST_K_DENSE:
            continue
        a, b = layer_span(i)
        if has_indexer(a, b) == want_full:
            pick = i; break
    if pick is None:  # fallback: any layer matching want_full
        for i in range(len(rms) - 1):
            a, b = layer_span(i)
            if has_indexer(a, b) == want_full:
                pick = i; break
    if pick is None:
        pick = FIRST_K_DENSE
    a, b = layer_span(pick)
    seq = []
    state = "MLA"
    for k in sorted([k for k in sk if a <= k["ts"] < b], key=lambda k: k["ts"]):
        sec, state = _canon_positional(k["name"], state)
        seq.append((sec, k["name"], round(timeavg.get(k["name"], 0.0), 2)))
    return pick, seq


BOLD = Font(name="Arial", bold=True); REG = Font(name="Arial")
HFILL = PatternFill("solid", fgColor="D9E1F2")
SF = {"MLA_attention": PatternFill("solid", fgColor="DDEBF7"),
      "MoE/MLP": PatternFill("solid", fgColor="E2EFDA"),
      "norm/comm": PatternFill("solid", fgColor="FCE4D6"),
      "other": PatternFill("solid", fgColor="F2F2F2")}
SECTIONS = ["MLA_attention", "MoE/MLP", "norm/comm", "other"]


def _analyze(atom_time, atom_struct_evs, sg_time, sg_struct_evs, want_full):
    ln, atom_seq = atom_layer(atom_struct_evs, want_full)
    sg_layer, sg_seq = sglang_layer_glm52(sg_struct_evs, sg_time, want_full)
    A = [(s, short(k), round(atom_time.get(k, 0.0), 2)) for s, k in atom_seq]
    S = [(s, short(k), v) for s, k, v in sg_seq]
    asum = collections.defaultdict(float); ssum = collections.defaultdict(float)
    for s, k, u in A: asum[s] += u
    for s, k, u in S: ssum[s] += u
    return ln, sg_layer, A, S, asum, ssum


BLOCK_W = 8   # columns per block: # | ATOM sec | ATOM kernel | ATOM us | spacer | SGLang sec | SGLang kernel | SGLang us
GAP = 3       # blank columns between the shared and full blocks


def _write_block(ws, c0, label, A, S, asum, ssum, body_rows):
    ws.cell(1, c0, label).font = BOLD
    hdr = ["#", "ATOM section", "ATOM kernel", "ATOM us", "", "SGLang section", "SGLang kernel", "SGLang us"]
    for j, h in enumerate(hdr):
        cell = ws.cell(2, c0 + j, h); cell.font = BOLD; cell.fill = HFILL; cell.alignment = Alignment(horizontal="center")
    r = 3
    for i in range(max(len(A), len(S))):
        asec, ak, au = A[i] if i < len(A) else ("", "", "")
        ssec, sk, su = S[i] if i < len(S) else ("", "", "")
        ws.cell(r, c0, i).font = REG
        for j, v in ((1, asec), (2, ak), (3, au)):
            cell = ws.cell(r, c0 + j, v if v != "" else None); cell.font = REG
            if j == 1 and asec in SF: cell.fill = SF[asec]
        for j, v in ((5, ssec), (6, sk), (7, su)):
            cell = ws.cell(r, c0 + j, v if v != "" else None); cell.font = REG
            if j == 5 and ssec in SF: cell.fill = SF[ssec]
        r += 1
    sr = body_rows + 4  # aligned subtotals across blocks
    ws.cell(sr, c0 + 1, "Section subtotals (us/layer)").font = BOLD; sr += 1
    ws.cell(sr, c0 + 1, "Section").font = BOLD; ws.cell(sr, c0 + 3, "ATOM").font = BOLD; ws.cell(sr, c0 + 7, "SGLang").font = BOLD; sr += 1
    for sec in SECTIONS:
        ws.cell(sr, c0 + 1, sec).font = REG
        if sec in SF: ws.cell(sr, c0 + 1).fill = SF[sec]
        ws.cell(sr, c0 + 3, round(asum.get(sec, 0.0), 1)).font = REG
        ws.cell(sr, c0 + 7, round(ssum.get(sec, 0.0), 1)).font = REG; sr += 1
    ws.cell(sr, c0 + 1, "TOTAL").font = BOLD
    ws.cell(sr, c0 + 3, round(sum(asum.values()), 1)).font = BOLD
    ws.cell(sr, c0 + 7, round(sum(ssum.values()), 1)).font = BOLD


def build(atom_time_p, atom_struct_p, sg_time_p, sg_struct_p, out_p, kinds, title):
    # load each (large) trace once, reuse for both shared & full
    atom_time = graph_name_avg(load(atom_time_p))
    atom_struct_evs = load(atom_struct_p)
    sg_time = _sglang_kernel_time_avg(load(sg_time_p))
    sg_struct_evs = load(sg_struct_p)

    order = [k for k in ("shared", "full") if k in kinds]
    data = {}
    for kind in order:
        ln, sgl, A, S, asum, ssum = _analyze(atom_time, atom_struct_evs, sg_time, sg_struct_evs, kind == "full")
        data[kind] = (ln, sgl, A, S, asum, ssum)
        print(f"[{kind}] ATOM layer {ln} | SGLang layer {sgl} | "
              + "ATOM " + ",".join(f"{s}={asum.get(s,0):.1f}" for s in SECTIONS)
              + " | SGLang " + ",".join(f"{s}={ssum.get(s,0):.1f}" for s in SECTIONS))

    body_rows = max(max(len(d[2]), len(d[3])) for d in data.values())
    wb = Workbook(); ws = wb.active; ws.title = "SideBySide"
    for bi, kind in enumerate(order):
        ln, sgl, A, S, asum, ssum = data[kind]
        c0 = 1 + bi * (BLOCK_W + GAP)
        _write_block(ws, c0, f"{kind.upper()}  (ATOM layer {ln} | SGLang layer {sgl})  —  {title}",
                     A, S, asum, ssum, body_rows)
    for bi in range(len(order)):
        c0 = 1 + bi * (BLOCK_W + GAP)
        for j, w in enumerate([5, 14, 46, 9, 3, 14, 46, 9]):
            ws.column_dimensions[chr(64 + c0 + j)].width = w
    ws.freeze_panes = "A3"
    wb.save(out_p); print("written", out_p)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--atom-time", required=True); ap.add_argument("--atom-struct", required=True)
    ap.add_argument("--sglang-time", required=True); ap.add_argument("--sglang-struct", required=True)
    ap.add_argument("--layer-kind", choices=["shared", "full", "both"], default="both",
                    help="both (default) -> one workbook with Summary + Shared + Full sheets")
    ap.add_argument("--out", required=True); ap.add_argument("--title", default="")
    a = ap.parse_args()
    kinds = ["shared", "full"] if a.layer_kind == "both" else [a.layer_kind]
    title = a.title or "GLM-5.2 MoE decode layer (ATOM | SGLang), CALL order"
    build(a.atom_time, a.atom_struct, a.sglang_time, a.sglang_struct, a.out, kinds, title)


if __name__ == "__main__":
    main()

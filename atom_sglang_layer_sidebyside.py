#!/usr/bin/env python3
"""
atom_sglang_layer_sidebyside.py

Build a side-by-side per-decode-layer comparison (SGLang vs ATOM), kernels in
CALL order, section-labeled (MLA_attention / MoE/MLP / norm/comm / other).

  ATOM order+labels : from --atom-struct (no-cuda-graph, gpu_user_annotations)
  ATOM timing       : from --atom-time (graph-ON decode[] step)
  SGLang            : from --sglang-step3 (analyze_trace.py step3 xlsx; already
                      per-layer, in execution order). Layer B (MoE) is used.

Usage:
  python atom_sglang_layer_sidebyside.py \
      --atom-time GRAPH.json.gz --atom-struct NOGRAPH.json.gz \
      --sglang-step3 step3_SGLang.xlsx --out out.xlsx
"""
import argparse, gzip, json, bisect, re
from collections import defaultdict, Counter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def load(path):
    with gzip.open(path, "rt") as f:
        t = json.load(f)
    return t if isinstance(t, list) else t.get("traceEvents", [])


def canon(sec, kn=""):
    # kernel-name override first: allreduce / rmsnorm are comm/norm regardless of
    # which module-section they were attributed to (e.g. cross_device_reduce that
    # SGLang emits inside the MoE section is really a TP allreduce = norm/comm).
    k = str(kn).lower()
    if ("cross_device_reduce" in k or "all_reduce" in k or "allreduce" in k
            or "nccl" in k or "reduce_1stage" in k):
        return "norm/comm"
    # q/k rmsnorm is MLA attention's internal q/k normalization (after q_a/kv_a
    # down-proj), NOT the residual-stream input/post-attn layernorm. Force it to
    # MLA_attention on both sides — ATOM annotates it under a generic "rmsnorm"
    # module (which the "norm" rule below would wrongly send to norm/comm), while
    # SGLang annotates it under DeepseekV2AttentionMLA.
    if "fused_qk_rmsnorm" in k or "qk_rmsnorm" in k:
        return "MLA_attention"
    # DSA sparse indexer kernels are part of attention (produce the topk selection);
    # classify by name so ATOM's (which land under generic annotations / "other")
    # and SGLang's line up in MLA_attention on both sides.
    if any(t in k for t in ("kn_entry_2c", "paged_mqa_logits", "topk_transform",
                            "radix_topk", "convert_req_index", "indexer_k_quant",
                            "hadamard", "wv_splitk")):
        return "MLA_attention"
    s = str(sec).lower()
    if s.startswith("prepare_"):
        return "norm/comm"
    if ("attention" in s or "attn" in s or "mla_decode" in s or "rope_and_kv" in s
            or "q_proj_and_k_up" in s or "v_up_proj_and_o" in s or "kv_cache" in s
            or "indexer" in s):
        return "MLA_attention"
    if any(t in s for t in ("moe", "mlp", "expert", "gate")):
        return "MoE/MLP"
    if "norm" in s or "nccl" in s or "allreduce" in s or "comm" in s:
        return "norm/comm"
    return "other"


def _k(evs):
    return [{"name": e.get("name","?"), "ts": float(e["ts"]), "dur": float(e.get("dur",0)),
             "pid": e.get("pid"), "tid": e.get("tid")}
            for e in evs if isinstance(e,dict) and e.get("ph")=="X"
            and str(e.get("cat","")).lower()=="kernel"]
def _g(evs):
    return [{"name": e.get("name","?"), "ts": float(e["ts"]), "dur": float(e.get("dur",0)),
             "pid": e.get("pid"), "tid": e.get("tid")}
            for e in evs if isinstance(e,dict) and e.get("ph")=="X"
            and e.get("cat")=="gpu_user_annotation"]
def _dom(ks): return Counter((k["pid"],k["tid"]) for k in ks).most_common(1)[0][0]


def graph_name_avg(evs):
    ks=_k(evs); p,t=_dom(ks); ks=[k for k in ks if (k["pid"],k["tid"])==(p,t)]
    ga=sorted([a for a in _g(evs) if a["name"].startswith("decode[")],key=lambda a:a["ts"])
    s=ga[len(ga)//2]; s0,s1=s["ts"],s["ts"]+s["dur"]
    agg=defaultdict(lambda:[0,0.0])
    for k in ks:
        if s0<=k["ts"]<s1: agg[k["name"]][0]+=1; agg[k["name"]][1]+=k["dur"]
    return {n:v[1]/v[0] for n,v in agg.items()}


def atom_layer(evs, want_full=False):
    ks=_k(evs); p,t=_dom(ks); ks=[k for k in ks if (k["pid"],k["tid"])==(p,t)]
    ga=[a for a in _g(evs) if (a["pid"],a["tid"])==(p,t)]
    l0=sorted([a for a in ga if a["name"].startswith("model.layers.0.")],key=lambda a:a["ts"])
    key=l0[0]["name"]; starts=sorted(a["ts"] for a in ga if a["name"]==key)
    segs=[(starts[i],starts[i+1]) for i in range(len(starts)-1)]
    kts=sorted(k["ts"] for k in ks); kd={k["ts"]:k["dur"] for k in ks}
    def sd(a,b):
        lo=bisect.bisect_left(kts,a);hi=bisect.bisect_left(kts,b);return sum(kd[x] for x in kts[lo:hi])
    durs=[sd(*s) for s in segs]
    f0,f1=segs[sorted(range(len(segs)),key=lambda i:durs[i])[len(segs)//2]]
    lay=defaultdict(list)
    for a in ga:
        if f0<=a["ts"]<f1:
            m=re.match(r"model\.layers\.(\d+)\.",a["name"])
            if m: lay[int(m.group(1))].append(a["ts"])
    layers=sorted(lay); lstart={n:min(lay[n]) for n in layers}
    def _has_indexer(s0,s1):
        # "full" indexer layer = runs the topk indexer this step (vs "shared" which
        # reuses a prior layer's selection). Detected by the indexer kernels.
        return any(("paged_mqa_logits" in k["name"] or "kn_entry_2c" in k["name"]
                    or "topk_transform" in k["name"]) and s0<=k["ts"]<s1 for k in ks)
    cands=[]
    for i,n in enumerate(layers):
        s0=lstart[n]; s1=lstart[layers[i+1]] if i+1<len(layers) else f1
        if n>=3 and any(a["name"]=="mxfp4_moe" and s0<=a["ts"]<s1 for a in ga):
            cands.append((n,s0,s1,_has_indexer(s0,s1)))
    chosen=None
    for (n,s0,s1,full) in cands:
        if full==want_full: chosen=(n,s0,s1); break
    if chosen is None and cands:
        n,s0,s1,_=cands[0]; chosen=(n,s0,s1)
    if chosen is None:
        n=layers[len(layers)//2]; i=layers.index(n)
        chosen=(n,lstart[n],lstart[layers[i+1]] if i+1<len(layers) else f1)
    n,s0,s1=chosen
    labs=sorted([a for a in ga if s0<=a["ts"]<s1 and not a["name"].startswith("decode")
                 and not a["name"].startswith("##")],key=lambda a:a["ts"])
    lts=[a["ts"] for a in labs]
    def inner(k):
        # smallest-duration annotation that *actually encloses* the kernel start
        # (a.ts <= k.ts < a.end). labs[:pos] all start at/before k.ts; keep only
        # those still open at k.ts, then take the innermost (smallest dur).
        pos=bisect.bisect_right(lts,k["ts"]);best=None;bd=1e18
        for j in range(pos-1,-1,-1):
            a=labs[j]
            if a["ts"]+a["dur"]<=k["ts"]: continue  # already closed -> not enclosing
            if a["dur"]<bd: best,bd=a,a["dur"]
        return best
    seq=[]
    for k in sorted([k for k in ks if s0<=k["ts"]<s1],key=lambda k:k["ts"]):
        a=inner(k); raw=a["name"] if a else "(unlabeled)"
        seq.append((canon(raw, k["name"]), k["name"]))
    return n, seq


def sglang_layer(path, want_full=False):
    wb=load_workbook(path,read_only=True,data_only=True);ws=wb.active
    hdr=[c.value for c in next(ws.iter_rows(min_row=1,max_row=1))];ix={h:i for i,h in enumerate(hdr)}
    rows=list(ws.iter_rows(min_row=2,values_only=True)); wb.close()
    # Auto-pick the MoE decode layer (matches the excel's "one MoE decode layer"):
    # the layer-type group whose kernels include an MoE GEMM (mfma_moe / moe_sorting).
    # Hardcoding "B" breaks across traces — e.g. the short-output (1024:16) no-graph
    # trace groups dense full-indexer layers as A/B and the MoE layer as C.
    layer_letters=[]
    for row in rows:
        lt=str(row[ix["LayerType"]] or "")
        if lt and lt[0].isalpha() and lt[1:2]==":" and lt[0] not in layer_letters:
            layer_letters.append(lt[0])
    def _rows(L): return [r for r in rows if str(r[ix["LayerType"]] or "").startswith(L+":")]
    def _has(L, *subs): return any(any(s in str(r[ix["KernelName"]] or "") for s in subs) for r in _rows(L))
    def _is_moe(L):  return _has(L, "mfma_moe", "moe_sorting")
    def _is_full(L): return _has(L, "paged_mqa_logits", "kn_entry_2c", "topk_transform")
    # Prefer an MoE layer matching the requested indexer kind (shared/full); else any
    # layer matching the kind (e.g. dense full-indexer if no full MoE layer captured);
    # else the first group.
    pick=None
    for L in layer_letters:
        if _is_moe(L) and _is_full(L)==want_full: pick=L; break
    if pick is None:
        for L in layer_letters:
            if _is_full(L)==want_full: pick=L; break
    if pick is None:
        pick=layer_letters[0] if layer_letters else "B"
    seq=[]
    for row in _rows(pick):
        kn=row[ix["KernelName"]]; sec=row[ix["Section"]]
        if not kn or sec=="Subtotal": continue
        av=row[ix["AvgDuration_us"]]
        try: av=float(av)
        except: av=0.0
        seq.append((canon(sec, kn), str(kn), av))
    return pick, seq


import subprocess, functools

@functools.lru_cache(maxsize=8192)
def _demangle(name):
    # Traces store raw C++ Itanium-mangled symbols for aiter template kernels
    # (e.g. _ZN5aiter30allreduce_fusion_kernel_1stageI...). Try a real demangler
    # first (llvm-cxxfilt handles the bf16 `DF16b` mangling; binutils c++filt is
    # often too old and returns it unchanged). If none works, fall back to a
    # version-independent heuristic that pulls the clean function name out of the
    # Itanium nested-name encoding (`_ZN <len><ns> <len><fn> ...`).
    if not name.startswith("_Z"):
        return name
    for tool in ("llvm-cxxfilt", "c++filt"):
        try:
            out = subprocess.run([tool, "--", name], capture_output=True, text=True, timeout=5)
            d = out.stdout.strip()
            if out.returncode == 0 and d and not d.startswith("_Z"):
                return d
        except Exception:
            pass
    m = re.match(r"_ZN(.*)", name)
    if m:
        s = m.group(1); idents = []; i = 0
        while i < len(s) and s[i].isdigit():
            j = i
            while j < len(s) and s[j].isdigit():
                j += 1
            ln = int(s[i:j]); idents.append(s[j:j + ln]); i = j + ln
        if idents:
            return idents[-1]  # innermost = function name (drops template/param mangling)
    return name


def short(n):
    n=_demangle(str(n))
    n=n.split("(")[0]
    for pre in ("void ","aiter::","_ZN5aiter","_ZN7sgl_hip","std::","__hip_","c10::"): n=n.replace(pre,"")
    return n[:52]


def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--atom-time",required=True); ap.add_argument("--atom-struct",required=True)
    ap.add_argument("--sglang-step3",required=True); ap.add_argument("--out",required=True)
    ap.add_argument("--title",default="")
    ap.add_argument("--layer-kind",choices=["shared","full"],default="shared",
                    help="shared = MoE decode layer that reuses the indexer selection "
                         "(default); full = a layer that runs the topk indexer this step.")
    a=ap.parse_args()
    want_full=(a.layer_kind=="full")
    atom_time=graph_name_avg(load(a.atom_time))
    ln,atom_seq=atom_layer(load(a.atom_struct), want_full)
    sg_letter,sg_seq=sglang_layer(a.sglang_step3, want_full)
    print(f"layer-kind={a.layer_kind}: ATOM layer {ln}, SGLang layer group {sg_letter}")

    A=[(s,short(k),round(atom_time.get(k,0.0),2)) for s,k in atom_seq]
    S=[(s,short(k),round(av,2)) for s,k,av in sg_seq]

    BOLD=Font(name="Arial",bold=True); REG=Font(name="Arial")
    HFILL=PatternFill("solid",fgColor="D9E1F2")
    SF={"MLA_attention":PatternFill("solid",fgColor="DDEBF7"),
        "MoE/MLP":PatternFill("solid",fgColor="E2EFDA"),
        "norm/comm":PatternFill("solid",fgColor="FCE4D6"),
        "other":PatternFill("solid",fgColor="F2F2F2")}
    wb=Workbook(); ws=wb.active; ws.title="SideBySide"
    ws.cell(1,1,a.title or f"{a.layer_kind} decode layer, kernels in CALL order (ATOM layer {ln} | SGLang group {sg_letter})").font=BOLD
    hdr=["#","ATOM section","ATOM kernel","ATOM us","","SGLang section","SGLang kernel","SGLang us"]
    for c,h in enumerate(hdr,1):
        cell=ws.cell(2,c,h); cell.font=BOLD; cell.fill=HFILL; cell.alignment=Alignment(horizontal="center")
    ws.freeze_panes="A3"; r=3
    for i in range(max(len(A),len(S))):
        asec,ak,au=A[i] if i<len(A) else ("","","")
        ssec,sk,su=S[i] if i<len(S) else ("","","")
        ws.cell(r,1,i).font=REG
        for c,v in ((2,asec),(3,ak),(4,au)):
            cell=ws.cell(r,c,v if v!="" else None); cell.font=REG
            if c==2 and asec in SF: cell.fill=SF[asec]
        for c,v in ((6,ssec),(7,sk),(8,su)):
            cell=ws.cell(r,c,v if v!="" else None); cell.font=REG
            if c==6 and ssec in SF: cell.fill=SF[ssec]
        r+=1
    r+=1; ws.cell(r,2,"Section subtotals (us/layer)").font=BOLD; r+=1
    asum=defaultdict(float); ssum=defaultdict(float)
    for s,k,u in A: asum[s]+=u
    for s,k,u in S: ssum[s]+=u
    ws.cell(r,2,"Section").font=BOLD; ws.cell(r,4,"ATOM").font=BOLD; ws.cell(r,8,"SGLang").font=BOLD; r+=1
    for sec in ["MLA_attention","MoE/MLP","norm/comm","other"]:
        ws.cell(r,2,sec).font=REG
        if sec in SF: ws.cell(r,2).fill=SF[sec]
        ws.cell(r,4,round(asum.get(sec,0.0),1)).font=REG
        ws.cell(r,8,round(ssum.get(sec,0.0),1)).font=REG; r+=1
    ws.cell(r,2,"TOTAL").font=BOLD; ws.cell(r,4,round(sum(asum.values()),1)).font=BOLD
    ws.cell(r,8,round(sum(ssum.values()),1)).font=BOLD
    for c,w in enumerate([5,14,52,9,3,14,52,9],1): ws.column_dimensions[chr(64+c)].width=w
    wb.save(a.out); print("written", a.out)
    print(f"ATOM layer {ln}: "+", ".join(f"{s}={asum.get(s,0):.1f}" for s in ("MLA_attention","MoE/MLP","norm/comm","other")))
    print("SGLang B: "+", ".join(f"{s}={ssum.get(s,0):.1f}" for s in ("MLA_attention","MoE/MLP","norm/comm","other")))


if __name__ == "__main__":
    main()

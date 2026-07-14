#!/usr/bin/env python3
from __future__ import annotations
"""
analyze_atom_trace.py

Decode-step kernel breakdown for ATOM (rocm/atom-dev) torch-profiler traces.

ATOM traces differ from SGLang traces: they have NO `nn.Module:` / `python_function`
events, so analyze_trace.py's layer detection does not apply. Instead ATOM's
`--mark-trace` emits `gpu_user_annotation` events on the SAME GPU stream as the
kernels:
  - `decode[bs=.. tok=.. d=..]`      → one whole decode step (forward pass)
  - `model.layers.N.self_attn.*`     → per-layer attention sub-modules
  - `model.layers.N.mlp.*`           → per-layer MLP/MoE sub-modules
  - `rmsnorm`, `rmsnorm_quant`, `kv_cache`, `mxfp4_moe`, ... → op-level regions

This tool isolates ONE decode step and assigns every GPU kernel to its innermost
enclosing annotation, then aggregates by (Section, LeafModule, KernelName). Output
matches analyze_trace.py's step3 schema so compare_breakdown.py can consume it.

Usage:
  python analyze_atom_trace.py --trace ATOM.trace.json.gz            # print
  python analyze_atom_trace.py --trace ATOM.trace.json.gz --out DIR --tag _ATOM
  python analyze_atom_trace.py --trace ATOM.trace.json.gz --step 5   # pick step index
"""
import argparse
import bisect
import gzip
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path


def load_trace(path: str):
    p = Path(path)
    if not p.exists():
        sys.exit(f"[ERROR] File not found: {path}")
    opener = gzip.open if path.endswith(".gz") else open
    with opener(path, "rt", encoding="utf-8") as f:
        return json.load(f)


def _events(trace):
    return trace if isinstance(trace, list) else trace.get("traceEvents", [])


def _is_x(e):
    return isinstance(e, dict) and e.get("ph") == "X"


def dominant_stream(kernels):
    """(pid, tid) carrying the most kernels."""
    c = Counter((k["pid"], k["tid"]) for k in kernels)
    return c.most_common(1)[0][0]


def fmt(us):
    if us >= 1e6:
        return f"{us/1e6:.3f} s"
    if us >= 1e3:
        return f"{us/1e3:.3f} ms"
    return f"{us:.1f} us"


# --- annotation name → (section, leaf) ------------------------------------
def classify(name: str):
    """Map a gpu_user_annotation name to (section, leaf)."""
    if name.startswith("model.layers."):
        parts = name.split(".")
        rest = parts[3:]            # drop model / layers / N
        if not rest:
            return "layer", ""
        section = rest[0]           # self_attn | mlp | ...
        leaf = ".".join(rest[1:])   # fused_qkv_a_proj | indexer.wq_b | gate ...
        return section, leaf
    if name.startswith("nccl"):
        return "comm", name.split(":", 1)[-1]
    # op-level regions (rmsnorm, rmsnorm_quant, kv_cache, mxfp4_moe, ...)
    return name, ""


# wrapper annotations that should never be a kernel's "innermost" label
def _is_wrapper(name: str) -> bool:
    return name.startswith("decode[") or name.startswith("## Call CompiledFxGraph")


def extract(trace):
    evs = _events(trace)
    kernels = [{"name": e.get("name", "?"), "ts": float(e["ts"]),
                "dur": float(e.get("dur", 0.0)), "pid": e.get("pid"), "tid": e.get("tid")}
               for e in evs if _is_x(e) and str(e.get("cat", "")).lower() == "kernel"]
    if not kernels:
        sys.exit("[ERROR] no kernel events in trace")
    gann = [{"name": e.get("name", "?"), "ts": float(e["ts"]),
             "dur": float(e.get("dur", 0.0)), "pid": e.get("pid"), "tid": e.get("tid")}
            for e in evs if _is_x(e) and e.get("cat") == "gpu_user_annotation"]
    return kernels, gann


def segment_forwards(gann):
    """Return (segments, method) where segments = [(ts0, ts1), ...] one per
    forward pass. Prefers the `decode[...]` wrapper (cuda-graph traces); falls
    back to the layer-0 first-submodule marker (no-cuda-graph traces, which have
    no decode[] wrapper but repeat model.layers.0.* every forward)."""
    dsteps = sorted([a for a in gann if a["name"].startswith("decode[")],
                     key=lambda a: a["ts"])
    if dsteps:
        return [(a["ts"], a["ts"] + a["dur"]) for a in dsteps], "decode[]"
    l0 = sorted([a for a in gann if a["name"].startswith("model.layers.0.")],
                key=lambda a: a["ts"])
    if not l0:
        sys.exit("[ERROR] no decode[] or model.layers.0.* annotations found.")
    key = l0[0]["name"]                      # forward-start boundary marker
    starts = sorted(a["ts"] for a in gann if a["name"] == key)
    segs = [(starts[i], starts[i + 1]) for i in range(len(starts) - 1)]
    return segs, key


def pick_decode_step(gann, kernels, step_idx):
    segs, method = segment_forwards(gann)
    if not segs:
        sys.exit("[ERROR] could not segment forward passes.")
    # per-segment kernel Σduration, to distinguish long prefill from short decode
    kts = sorted(k["ts"] for k in kernels)
    kdur = {k["ts"]: k["dur"] for k in kernels}
    def seg_dur(s0, s1):
        lo = bisect.bisect_left(kts, s0); hi = bisect.bisect_left(kts, s1)
        return sum(kdur[t] for t in kts[lo:hi])
    durs = [seg_dur(*s) for s in segs]
    print(f"[INFO] segmentation by '{method}': {len(segs)} forward passes; "
          f"kernel-Σdur min/median/max = "
          f"{fmt(min(durs))}/{fmt(sorted(durs)[len(durs)//2])}/{fmt(max(durs))}",
          file=sys.stderr)
    if step_idx is None:
        # steady-state decode ~ the median-duration segment (prefill is the long outlier)
        step_idx = sorted(range(len(segs)), key=lambda i: durs[i])[len(segs) // 2]
    step_idx = max(0, min(step_idx, len(segs) - 1))
    s0, s1 = segs[step_idx]
    print(f"[INFO] using forward #{step_idx} (kernel Σdur {fmt(durs[step_idx])})",
          file=sys.stderr)
    return {"ts": s0, "dur": s1 - s0}


def innermost_label(k, ann_sorted, ann_ts):
    """Smallest-duration non-wrapper annotation whose span contains kernel start."""
    ts = k["ts"]
    pos = bisect.bisect_right(ann_ts, ts)
    best = None
    best_dur = float("inf")
    # scan backwards over annotations that started at/before the kernel
    for j in range(pos - 1, -1, -1):
        a = ann_sorted[j]
        if a["ts"] + a["dur"] < ts:
            continue
        if a["dur"] < best_dur:
            best, best_dur = a, a["dur"]
        # annotations are short; a few thousand us back is plenty
        if a["ts"] < ts - 200_000:
            break
    return best


def _one_step(trace, step_idx):
    """Return (kernels_in_step, step_span, gann_in_step) for one decode forward."""
    kernels, gann = extract(trace)
    pid, tid = dominant_stream(kernels)
    kernels = [k for k in kernels if (k["pid"], k["tid"]) == (pid, tid)]
    gann = [a for a in gann if (a["pid"], a["tid"]) == (pid, tid)]
    step = pick_decode_step(gann, kernels, step_idx)
    s0, s1 = step["ts"], step["ts"] + step["dur"]
    kin = [k for k in kernels if s0 <= k["ts"] < s1]
    gin = sorted([a for a in gann if s0 <= a["ts"] < s1 and not _is_wrapper(a["name"])],
                 key=lambda a: a["ts"])
    return kin, (s0, s1), gin


def build_struct_map(struct_trace, step_idx):
    """kernel_name -> (section, leaf), from one decode forward of a (no-cuda-graph)
    trace whose annotations wrap the kernels. Majority label per kernel name."""
    kin, _, gin = _one_step(struct_trace, step_idx)
    lab_ts = [a["ts"] for a in gin]
    votes = defaultdict(Counter)
    for k in kin:
        a = innermost_label(k, gin, lab_ts)
        lbl = classify(a["name"]) if a else ("(unlabeled)", "")
        votes[k["name"]][lbl] += 1
    return {name: c.most_common(1)[0][0] for name, c in votes.items()}


def union_busy(kernels):
    """Wall time the GPU is busy (overlaps counted once), in us."""
    iv = sorted((k["ts"], k["ts"] + k["dur"]) for k in kernels)
    if not iv:
        return 0.0
    tot = 0.0; cs, ce = iv[0]
    for s, e in iv[1:]:
        if s > ce:
            tot += ce - cs; cs, ce = s, e
        else:
            ce = max(ce, e)
    return tot + (ce - cs)


def analyze(time_trace, struct_map, step_idx=None):
    """Timing from the graph-ON trace (accurate); labels from struct_map."""
    kin, (s0, s1), _ = _one_step(time_trace, step_idx)
    agg = defaultdict(lambda: {"count": 0, "sum": 0.0})
    order = []
    for k in kin:
        section, leaf = struct_map.get(k["name"], ("(unlabeled)", ""))
        key = (section, leaf, k["name"])
        if key not in agg:
            order.append(key)
        agg[key]["count"] += 1
        agg[key]["sum"] += k["dur"]
    total = sum(v["sum"] for v in agg.values())
    rows = []
    for key in order:
        section, leaf, kname = key
        v = agg[key]
        rows.append({"section": section, "leaf": leaf, "kernel": kname,
                     "count": v["count"], "sum": v["sum"],
                     "avg": v["sum"] / v["count"], "pct": v["sum"] / total * 100})
    unlabeled = sum(v["sum"] for k, v in agg.items() if k[0] == "(unlabeled)")
    busy = union_busy(kin)
    return rows, total, s1 - s0, unlabeled, busy


# --- section-level rollup --------------------------------------------------
def section_rollup(rows):
    sec = defaultdict(lambda: {"sum": 0.0, "count": 0})
    for r in rows:
        sec[r["section"]]["sum"] += r["sum"]
        sec[r["section"]]["count"] += r["count"]
    return sorted(sec.items(), key=lambda kv: kv[1]["sum"], reverse=True)


def print_report(rows, total, step_dur, unlabeled, busy):
    print(f"\n{'='*96}")
    print(" ATOM decode-step kernel breakdown (timing: graph-ON, labels: no-cuda-graph)")
    print(f" decode-step GPU span: {fmt(step_dur)} | GPU-busy: {fmt(busy)}"
          f" | kernel Σduration: {fmt(total)} | unlabeled: {fmt(unlabeled)}")
    print(f"{'='*96}")
    print("\n-- Section rollup (Σ kernel duration in one decode step) --")
    print(f"  {'Section':<26} {'Σdur':>12} {'Kernels':>8} {'Pct':>6}")
    print(f"  {'-'*26} {'-'*12} {'-'*8} {'-'*6}")
    for name, v in section_rollup(rows):
        print(f"  {name:<26} {fmt(v['sum']):>12} {v['count']:>8} "
              f"{v['sum']/total*100:>5.1f}%")

    print("\n-- Per (section, leaf, kernel) --")
    print(f"  {'Section':<16} {'Leaf':<22} {'Avg':>9} {'Cnt':>4} {'Σdur':>10} {'Pct':>5}  Kernel")
    print(f"  {'-'*16} {'-'*22} {'-'*9} {'-'*4} {'-'*10} {'-'*5}  {'-'*40}")
    for r in sorted(rows, key=lambda r: r["sum"], reverse=True):
        print(f"  {r['section']:<16} {(r['leaf'] or '-'):<22} {fmt(r['avg']):>9} "
              f"{r['count']:>4} {fmt(r['sum']):>10} {r['pct']:>4.1f}%  {r['kernel'][:40]}")
    print(f"{'='*96}\n")


def write_step3_xlsx(rows, path):
    """Write analyze_trace.py step3-compatible schema so compare_breakdown.py works."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    headers = ["LayerType", "LayerCount", "Index", "Section", "LeafModule",
               "KernelName", "AvgDuration_us", "Count", "SumDuration_us",
               "Percentage", "MatchMethod", "GraphOFF_KernelName",
               "GraphOFF_Duration_us", "CallSite"]
    wb = Workbook(); ws = wb.active
    bold = Font(name="Arial", size=10, bold=True); reg = Font(name="Arial", size=10)
    fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = bold; cell.fill = fill; cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    r = 2
    for i, row in enumerate(sorted(rows, key=lambda r: (r["section"], r["leaf"]))):
        vals = ["decode: ATOM", 1, i, row["section"], row["leaf"] or "(self)",
                row["kernel"], round(row["avg"], 3), row["count"],
                round(row["sum"], 1), round(row["pct"], 2), "atom-annotation",
                row["kernel"], round(row["avg"], 3), ""]
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=v).font = reg
        r += 1
    for c in range(1, len(headers) + 1):
        w = len(headers[c - 1])
        for rr in range(2, min(r, 60)):
            v = ws.cell(row=rr, column=c).value
            if v is not None:
                w = max(w, min(len(str(v)), 60))
        ws.column_dimensions[get_column_letter(c)].width = w + 2
    wb.save(path)
    print(f"[INFO] step3 breakdown written to {path}", file=sys.stderr)


def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--time-trace", required=True, metavar="TRACE",
                   help="graph-ON ATOM trace (accurate kernel timing)")
    p.add_argument("--struct-trace", metavar="TRACE",
                   help="no-cuda-graph ATOM trace (annotation labels). "
                        "If omitted, kernels are left unlabeled.")
    p.add_argument("--step", type=int, default=None,
                   help="decode step/forward index (default: median-duration one)")
    p.add_argument("--out", metavar="DIR", help="write step3_layer_breakdown xlsx here")
    p.add_argument("--tag", default="_ATOM", help="filename tag (default: _ATOM)")
    args = p.parse_args()

    struct_map = {}
    if args.struct_trace:
        print(f"[INFO] structure trace: {args.struct_trace}", file=sys.stderr)
        struct_map = build_struct_map(load_trace(args.struct_trace), args.step)

    print(f"[INFO] timing trace: {args.time_trace}", file=sys.stderr)
    rows, total, step_dur, unlabeled, busy = analyze(
        load_trace(args.time_trace), struct_map, args.step)
    print_report(rows, total, step_dur, unlabeled, busy)
    if args.out:
        out = Path(args.out); out.mkdir(parents=True, exist_ok=True)
        write_step3_xlsx(rows, str(out / f"step3_layer_breakdown{args.tag}.xlsx"))


if __name__ == "__main__":
    main()

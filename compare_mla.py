#!/usr/bin/env python3
"""
compare_mla.py

Canonical per-decode-step section comparison of SGLang vs ATOM (MI355X), so the
MLA attention block (SGLang `DeepseekV2AttentionMLA` vs ATOM `model.layers.*.self_attn`
+ MLA sub-op annotations) is directly comparable, despite ATOM torch.compiling the
model (which flattens the nn.Module hierarchy, so there is no `DeepseekV2MLAAttention`
module event to match SGLang's).

Inputs:
  --sg-time   SGLang graph-ON decode trace (.json.gz)     accurate per-kernel time
  --sg-step3  SGLang step3 xlsx (from analyze_trace.py)   kernel-name -> section
  --atom-step3 ATOM step3 xlsx (from analyze_atom_trace.py) already per-decode-step

SGLang per-step time = Σ(kernel global duration) / n_decode_steps, where n_decode_steps
is inferred from the per-step 'topk_transform_decode' kernel count // 78 layers.
"""
import argparse
import gzip
import json
import sys
from collections import Counter, defaultdict
from openpyxl import load_workbook

N_LAYERS = 78  # GLM-5.x decoder layers


def canon(section: str) -> str:
    s = section.lower()
    if any(t in s for t in ("attn", "attention", "mla_decode", "rope_and_kv",
                             "q_proj_and_k_up", "v_up_proj_and_o", "kv_cache")):
        return "MLA_attention"
    if any(t in s for t in ("moe", "mlp", "expert")):
        return "MoE/MLP"
    if "norm" in s:
        return "norm"
    if "comm" in s or "nccl" in s or "allreduce" in s:
        return "comm"
    return "other"


def read_step3_map(path):
    """kernel_name -> section (for SGLang, to label graph-on kernels)."""
    wb = load_workbook(path, read_only=True, data_only=True); ws = wb.active
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ix = {h: i for i, h in enumerate(hdr)}
    m = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        kn = row[ix.get("KernelName")]; sec = row[ix.get("Section")]
        if kn and sec and sec != "Subtotal":
            m.setdefault(kn, sec)
    wb.close()
    return m


def read_atom_step3_perstep(path):
    """canonical -> per-step us, from ATOM step3 (SumDuration_us is per decode step)."""
    wb = load_workbook(path, read_only=True, data_only=True); ws = wb.active
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ix = {h: i for i, h in enumerate(hdr)}
    out = defaultdict(float)
    for row in ws.iter_rows(min_row=2, values_only=True):
        sec = row[ix.get("Section")]; s = row[ix.get("SumDuration_us")]
        if sec and sec != "Subtotal" and s not in (None, ""):
            out[canon(str(sec))] += float(s)
    wb.close()
    return out


def sglang_perstep(time_trace, step3_map):
    with gzip.open(time_trace, "rt") as f:
        t = json.load(f)
    evs = t if isinstance(t, list) else t.get("traceEvents", [])
    k = [e for e in evs if isinstance(e, dict)
         and str(e.get("cat", "")).lower() == "kernel" and e.get("ph") == "X"]
    name_sum = defaultdict(float)
    for e in k:
        name_sum[e.get("name", "?")] += float(e.get("dur", 0.0))
    # decode steps: topk_transform_decode fires once per layer per step
    tt = sum(c for n, c in Counter(e.get("name", "") for e in k).items()
             if "topk_transform_decode" in n)
    n_steps = max(1, tt // N_LAYERS)
    out = defaultdict(float)
    unlabeled = 0.0
    for name, s in name_sum.items():
        sec = step3_map.get(name)
        if sec is None:
            unlabeled += s; out["other"] += s / n_steps
        else:
            out[canon(sec)] += s / n_steps
    return out, n_steps


def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--sg-time", required=True)
    p.add_argument("--sg-step3", required=True)
    p.add_argument("--atom-step3", required=True)
    args = p.parse_args()

    sg_map = read_step3_map(args.sg_step3)
    sg, n_steps = sglang_perstep(args.sg_time, sg_map)
    at = read_atom_step3_perstep(args.atom_step3)

    print(f"[INFO] SGLang decode steps inferred: {n_steps}", file=sys.stderr)
    buckets = ["MLA_attention", "MoE/MLP", "norm", "comm", "other"]
    print(f"\n{'='*64}")
    print(" Per-decode-step GPU kernel time (Σduration, us)  —  MI355X TP4")
    print(f"{'='*64}")
    print(f"  {'Section':<16} {'SGLang':>12} {'ATOM':>12} {'ATOM/SGLang':>12}")
    print(f"  {'-'*16} {'-'*12} {'-'*12} {'-'*12}")
    tsg = tat = 0.0
    for b in buckets:
        a, s = at.get(b, 0.0), sg.get(b, 0.0)
        tsg += s; tat += a
        ratio = f"{a/s*100:.0f}%" if s else "-"
        print(f"  {b:<16} {s:>12.1f} {a:>12.1f} {ratio:>12}")
    print(f"  {'-'*16} {'-'*12} {'-'*12} {'-'*12}")
    print(f"  {'TOTAL':<16} {tsg:>12.1f} {tat:>12.1f} "
          f"{(tat/tsg*100 if tsg else 0):>11.0f}%")
    print(f"{'='*64}\n")


if __name__ == "__main__":
    main()

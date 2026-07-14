#!/usr/bin/env python3
"""Build a decode-phase kernel comparison xlsx (MI355X mxfp4 vs B200 nvfp4).

Decode runs with cuda-graph ON so there is no per-layer nn.Module structure;
we therefore compare by kernel functional category + full per-kernel lists,
all normalised to per decode step.
"""
import gzip, json, sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

NSTEPS = 5  # decode steps captured in the profile (topk_transform calls / 78 layers)

CAT_ORDER = ["Attention", "GEMM (dense)", "MoE (GEMM+route)", "Comm/AllReduce",
             "Quant", "Norm/Elem", "Other"]


def classify(n: str) -> str:
    s = n.lower()
    if any(k in s for k in ['allreduce', 'quickreduce', 'cross_device_reduce',
                            'all_reduce', 'oneshot', 'twoshot', 'reduce_fusion']):
        return 'Comm/AllReduce'
    if any(k in s for k in ['moe', 'expert', 'moe_sorting']):
        return 'MoE (GEMM+route)'
    if any(k in s for k in ['topk', 'biased_grouped', 'grouped_topk', 'routing']):
        return 'MoE (GEMM+route)'
    if any(k in s for k in ['fmha', 'mla', 'main_kernel', 'paged', '_attn',
                            'flash', 'cached_indirect', 'rope', 'attention']):
        return 'Attention'
    if any(k in s for k in ['quantize', 'dynamic_per', '_quant', 'scaled_quant']):
        return 'Quant'
    if any(k in s for k in ['gemm', 'hgemm', 'nvjet', 'cublas', 'bmm', 'cutlass',
                            'matmul', 'splitkreduce', 'wgmma', 'nvfp4', 'e2m1']):
        return 'GEMM (dense)'
    if any(k in s for k in ['norm', 'silu', 'add', 'elementwise', 'vectorized',
                            'activation']):
        return 'Norm/Elem'
    return 'Other'


def load(path: str) -> dict:
    t = json.load(gzip.open(path, 'rt'))
    evs = t if isinstance(t, list) else t.get("traceEvents", [])
    ker = [e for e in evs if isinstance(e, dict)
           and str(e.get("cat", "")).lower() == "kernel" and e.get("ph") == "X"]
    gmin = min(e["ts"] for e in ker)
    gmax = max(e["ts"] + e["dur"] for e in ker)
    iv = sorted((e["ts"], e["ts"] + e["dur"]) for e in ker)
    ub = 0.0
    cs, ce = iv[0]
    for s, e in iv[1:]:
        if s > ce:
            ub += ce - cs
            cs, ce = s, e
        else:
            ce = max(ce, e)
    ub += ce - cs
    agg = defaultdict(lambda: [0, 0.0])
    kern = defaultdict(lambda: [0, 0.0])
    for e in ker:
        c = classify(e["name"])
        agg[c][0] += 1
        agg[c][1] += e["dur"]
        kern[e["name"]][0] += 1
        kern[e["name"]][1] += e["dur"]
    return dict(wall=gmax - gmin, busy=ub, sumd=sum(e["dur"] for e in ker),
                streams=len(set(e["tid"] for e in ker)), nk=len(ker),
                agg={k: v for k, v in agg.items()},
                kern={k: v for k, v in kern.items()})


ARIAL = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
HFILL = PatternFill("solid", fgColor="D9E1F2")
SUMFILL = PatternFill("solid", fgColor="FFF2CC")
BORDER = Border(*([Side(style="thin", color="CCCCCC")] * 4))


def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD
        cell.fill = HFILL
        cell.alignment = Alignment(horizontal="center")


def autosize(ws, ncol, maxrow=80):
    for c in range(1, ncol + 1):
        m = 8
        for r in range(1, min(ws.max_row, maxrow) + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                m = max(m, min(len(str(v)), 70))
        ws.column_dimensions[get_column_letter(c)].width = m + 2


def main():
    mx_path, nv_path, out = sys.argv[1], sys.argv[2], sys.argv[3]
    mx, nv = load(mx_path), load(nv_path)

    wb = Workbook()
    # ---------- Summary sheet ----------
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value="GLM-5 FP4 DECODE comparison (per decode step, TP-0, in8192/out1024/conc4)").font = BOLD
    ws.cell(row=2, column=1, value=f"MI355X = mxfp4 (ROCm) | B200 = nvfp4 (CUDA) | {NSTEPS} steps profiled").font = ARIAL

    # Top-line metrics table
    r = 4
    hdr = ["Metric", "MI355X (mxfp4)", "B200 (nvfp4)", "B200/MI355X"]
    for c, h in enumerate(hdr, 1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, len(hdr))
    metrics = [
        ("Wall time / step (ms)", mx['wall'] / NSTEPS / 1e3, nv['wall'] / NSTEPS / 1e3, True),
        ("GPU busy (union) / step (ms)", mx['busy'] / NSTEPS / 1e3, nv['busy'] / NSTEPS / 1e3, True),
        ("Sum kernel dur / step (ms)", mx['sumd'] / NSTEPS / 1e3, nv['sumd'] / NSTEPS / 1e3, True),
        ("Kernel overlap factor (sum/busy)", mx['sumd'] / mx['busy'], nv['sumd'] / nv['busy'], False),
        ("GPU streams used", mx['streams'], nv['streams'], False),
        ("Kernel launches / step", mx['nk'] / NSTEPS, nv['nk'] / NSTEPS, False),
    ]
    r += 1
    for name, a, b, ratio in metrics:
        ws.cell(row=r, column=1, value=name).font = ARIAL
        ca = ws.cell(row=r, column=2, value=round(a, 2)); ca.font = ARIAL
        cb = ws.cell(row=r, column=3, value=round(b, 2)); cb.font = ARIAL
        cr = ws.cell(row=r, column=4, value=(round(b / a, 3) if (ratio and a) else ""))
        cr.font = ARIAL
        if ratio:
            cr.number_format = "0%"
        r += 1

    # speed note
    ws.cell(row=r + 0, column=1,
            value=f"=> MI355X decode throughput = {nv['wall']/mx['wall']*100:.0f}% of B200 "
                  f"(per-step wall {mx['wall']/NSTEPS/1e3:.2f} vs {nv['wall']/NSTEPS/1e3:.2f} ms)").font = BOLD

    # Category table
    r += 3
    ws.cell(row=r, column=1, value="Per-category GPU work (sum of kernel dur per step, ms)").font = BOLD
    r += 1
    chdr = ["Category", "MI355X ms", "MI355X %", "B200 ms", "B200 %", "B200/MI355X"]
    for c, h in enumerate(chdr, 1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, len(chdr))
    r += 1
    mx_tot = mx['sumd'] / NSTEPS / 1e3
    nv_tot = nv['sumd'] / NSTEPS / 1e3
    cats = [c for c in CAT_ORDER if c in mx['agg'] or c in nv['agg']]
    for c in cats:
        a = mx['agg'].get(c, [0, 0.0])[1] / NSTEPS / 1e3
        b = nv['agg'].get(c, [0, 0.0])[1] / NSTEPS / 1e3
        ws.cell(row=r, column=1, value=c).font = ARIAL
        ws.cell(row=r, column=2, value=round(a, 2)).font = ARIAL
        ws.cell(row=r, column=3, value=round(a / mx_tot * 100, 1)).font = ARIAL
        ws.cell(row=r, column=4, value=round(b, 2)).font = ARIAL
        ws.cell(row=r, column=5, value=round(b / nv_tot * 100, 1)).font = ARIAL
        cr = ws.cell(row=r, column=6, value=(round(b / a, 3) if a else ""))
        cr.font = ARIAL
        if a:
            cr.number_format = "0%"
        r += 1
    # total row
    ws.cell(row=r, column=1, value="TOTAL (sum dur)").font = BOLD
    ws.cell(row=r, column=2, value=round(mx_tot, 2)).font = BOLD
    ws.cell(row=r, column=4, value=round(nv_tot, 2)).font = BOLD
    cr = ws.cell(row=r, column=6, value=round(nv_tot / mx_tot, 3)); cr.font = BOLD
    cr.number_format = "0%"
    for rr in range(4, r + 1):
        for cc in range(1, 7):
            ws.cell(row=rr, column=cc).border = BORDER
    autosize(ws, 6)
    ws.freeze_panes = "A1"

    # ---------- per-platform kernel sheets ----------
    def kernel_sheet(title, data):
        s = wb.create_sheet(title)
        hdr = ["Category", "KernelName", "Calls/step", "Dur/step (ms)",
               "Avg (us)", "% of decode"]
        for c, h in enumerate(hdr, 1):
            s.cell(row=1, column=c, value=h)
        style_header(s, 1, len(hdr))
        tot = data['sumd']
        rows = sorted(data['kern'].items(), key=lambda kv: -kv[1][1])
        r = 2
        for name, (cnt, dur) in rows:
            s.cell(row=r, column=1, value=classify(name)).font = ARIAL
            s.cell(row=r, column=2, value=name).font = ARIAL
            s.cell(row=r, column=3, value=round(cnt / NSTEPS, 1)).font = ARIAL
            s.cell(row=r, column=4, value=round(dur / NSTEPS / 1e3, 3)).font = ARIAL
            s.cell(row=r, column=5, value=round(dur / cnt, 2)).font = ARIAL
            s.cell(row=r, column=6, value=round(dur / tot * 100, 2)).font = ARIAL
            r += 1
        autosize(s, len(hdr))
        s.freeze_panes = "A2"

    kernel_sheet("MI355X_kernels", mx)
    kernel_sheet("B200_kernels", nv)

    wb.save(out)
    print(f"[INFO] written {out}")
    print(f"MI355X wall/step {mx['wall']/NSTEPS/1e3:.2f}ms  B200 wall/step {nv['wall']/NSTEPS/1e3:.2f}ms  "
          f"=> MI355X = {nv['wall']/mx['wall']*100:.0f}% of B200")


if __name__ == "__main__":
    main()
